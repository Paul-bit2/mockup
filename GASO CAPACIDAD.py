"""
GASO COMUNICACIONES – IN-OUT Report Processor
Cleans, enriches and transforms the IN-OUT Excel database and generates
an executive client report.

Persistent decision memory: manual XDOCK assignments are saved to
'gaso_decisions.json' so they auto-apply on future uploads.
"""

import io
import json
import os
import re
import datetime
import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from unidecode import unidecode
import base64
import tempfile
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Image as RLImage,
                                 Table, TableStyle, PageBreak, HRFlowable, KeepTogether)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY

# ─────────────────────────────────────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
DECISIONS_FILE = "gaso_decisions.json"   # persisted next to the script
LOGO_PATH = "GASO_COMUNICACIONES_LOGO.jpg"  # place beside the script

GASO_BLUE   = "#1A3A6B"
GASO_LIGHT  = "#2E6DB4"
GASO_ACCENT = "#4A90D9"

CAPACIDADES = {
    "Gaso- Tijuana-E-NS":      677,
    "Gaso- La Paz-E-NS":       316,
    "Gaso- Hermosillo-E-NS":   400,
    "Gaso- Culiacán-E-NS":     600,
    "Gaso- Cd. Juarez-E-NS":   350,
    "Gaso- Chihuahua-E-NS":    400,
    "Gaso- Monterrey-E-NS":   1500,
    "Gaso- Guadalajara-E-NS": 1200,
    "Gaso- Querétaro-E-NS":    800,
    "Gaso- Torreón-E-NS":      250,
}

REGION_MAP = {
    "Gaso- La Paz-E-NS":       "REGIÓN JOSÉ",
    "Gaso- Culiacán-E-NS":     "REGIÓN JOSÉ",
    "Gaso- Guadalajara-E-NS":  "REGIÓN JOSÉ",
    "Gaso- Querétaro-E-NS":    "REGIÓN JOSÉ",
    "Gaso- Tijuana-E-NS":      "REGIÓN JORGE",
    "Gaso- Hermosillo-E-NS":   "REGIÓN JORGE",
    "Gaso- Cd. Juarez-E-NS":   "REGIÓN JORGE",
    "Gaso- Chihuahua-E-NS":    "REGIÓN JORGE",
    "Gaso- Monterrey-E-NS":    "REGIÓN JORGE",
    "Gaso- Torreón-E-NS":      "REGIÓN JORGE",
}

CIUDAD_MAP = {
    "Gaso- Tijuana-E-NS":      "Tijuana",
    "Gaso- La Paz-E-NS":       "La Paz",
    "Gaso- Hermosillo-E-NS":   "Hermosillo",
    "Gaso- Culiacán-E-NS":     "Culiacán",
    "Gaso- Cd. Juarez-E-NS":   "Cd. Juárez",
    "Gaso- Chihuahua-E-NS":    "Chihuahua",
    "Gaso- Monterrey-E-NS":    "Monterrey",
    "Gaso- Guadalajara-E-NS":  "Guadalajara",
    "Gaso- Querétaro-E-NS":    "Querétaro",
    "Gaso- Torreón-E-NS":      "Torreón",
}

M2_TIPO = {
    "EUROPALLET":        0.96,
    "ESTANDAR":          1.44,
    "SOBREDIMENSIONADA": 3.66,
}

FOLIO_XDOCK = [
    ("GTOR",  "Gaso- Torreón-E-NS"),
    ("GCCJS", "Gaso- Cd. Juarez-E-NS"),
    ("GASO",  "Gaso- Chihuahua-E-NS"),
    ("CUL",   "Gaso- Culiacán-E-NS"),
    ("GCH",   "Gaso- Hermosillo-E-NS"),
    ("BS",    "Gaso- La Paz-E-NS"),
    ("GCM",   "Gaso- Monterrey-E-NS"),
    ("GCQ",   "Gaso- Querétaro-E-NS"),
    ("GCTIJ", "Gaso- Tijuana-E-NS"),
]

ID_SITIO_PREFIXES = {
    "JAL": "Gaso- Guadalajara-E-NS",
    "SIN": "Gaso- Culiacán-E-NS",
    "NL":  "Gaso- Monterrey-E-NS",
    "NLE": "Gaso- Monterrey-E-NS",
}

KEYWORD_XDOCK = {
    "culiacan":    "Gaso- Culiacán-E-NS",
    "culiacán":    "Gaso- Culiacán-E-NS",
    "gdl":         "Gaso- Guadalajara-E-NS",
    "guadalajara": "Gaso- Guadalajara-E-NS",
    "mty":         "Gaso- Monterrey-E-NS",
    "monterrey":   "Gaso- Monterrey-E-NS",
    "qro":         "Gaso- Querétaro-E-NS",
    "queretaro":   "Gaso- Querétaro-E-NS",
    "querétaro":   "Gaso- Querétaro-E-NS",
    "hermosillo":  "Gaso- Hermosillo-E-NS",
    "juarez":      "Gaso- Cd. Juarez-E-NS",
    "juárez":      "Gaso- Cd. Juarez-E-NS",
    "chihuahua":   "Gaso- Chihuahua-E-NS",
    "torreon":     "Gaso- Torreón-E-NS",
    "torreón":     "Gaso- Torreón-E-NS",
    "tijuana":     "Gaso- Tijuana-E-NS",
    "la paz":      "Gaso- La Paz-E-NS",
    "lapaz":       "Gaso- La Paz-E-NS",
}

PALLET_NORM = {
    "antena gd":         "SOBREDIMENSIONADA",
    "antena md":         "ESTANDAR",
    "antena ch":         "EUROPALLET",
    "gabinete gd":       "ESTANDAR",
    "gabinete med":      "ESTANDAR",
    "galvanizado":       "SOBREDIMENSIONADA",
    "soporte":           "SOBREDIMENSIONADA",
    "estandar":          "ESTANDAR",
    "europallet":        "EUROPALLET",
    "sobredimensionado": "SOBREDIMENSIONADA",
    "sobredimensionada": "SOBREDIMENSIONADA",
}

XDOCK_ALIASES = {
    "gaso- torreon-e-ns":     "Gaso- Torreón-E-NS",
    "gaso- torreón- e-ns":    "Gaso- Torreón-E-NS",
    "gaso- torreon- e-ns":    "Gaso- Torreón-E-NS",
    "gaso- tijuana-e-ns":     "Gaso- Tijuana-E-NS",
    "gaso- la paz-e-ns":      "Gaso- La Paz-E-NS",
    "gaso- hermosillo-e-ns":  "Gaso- Hermosillo-E-NS",
    "gaso- culiacan-e-ns":    "Gaso- Culiacán-E-NS",
    "gaso- culiacán-e-ns":    "Gaso- Culiacán-E-NS",
    "gaso- cd. juarez-e-ns":  "Gaso- Cd. Juarez-E-NS",
    "gaso- chihuahua-e-ns":   "Gaso- Chihuahua-E-NS",
    "gaso- monterrey-e-ns":   "Gaso- Monterrey-E-NS",
    "gaso- guadalajara-e-ns": "Gaso- Guadalajara-E-NS",
    "gaso- queretaro-e-ns":   "Gaso- Querétaro-E-NS",
    "gaso- querétaro-e-ns":   "Gaso- Querétaro-E-NS",
}

XDOCK_OPTIONS = ["— Selecciona —"] + sorted(CAPACIDADES.keys())

# ─────────────────────────────────────────────────────────────────────────────
#  PERSISTENT DECISIONS  (saved to gaso_decisions.json beside the script)
# ─────────────────────────────────────────────────────────────────────────────
def load_decisions() -> dict:
    """Load saved ID_SITIO → {action, xdock} decisions from disk."""
    if os.path.exists(DECISIONS_FILE):
        try:
            with open(DECISIONS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_decisions(decisions: dict):
    """Persist decisions to disk."""
    with open(DECISIONS_FILE, "w", encoding="utf-8") as f:
        json.dump(decisions, f, ensure_ascii=False, indent=2)


def delete_decision(id_sitio: str):
    d = load_decisions()
    if id_sitio in d:
        del d[id_sitio]
        save_decisions(d)


# ─────────────────────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def norm(s):
    if s is None:
        return ""
    return unidecode(str(s)).strip().lower()


def is_empty_xdock(val):
    """True if XDOCK is blank/null — candidate for inference."""
    n = norm(val)
    return not n or n in ("none", "nan")

def is_selecciona(val):
    """True if XDOCK is a 'Selecciona...' placeholder — always delete."""
    return norm(val).startswith("seleccion")


def contains_ens(val):
    return "e-ns" in norm(val)


# ─────────────────────────────────────────────────────────────────────────────
#  LOAD
# ─────────────────────────────────────────────────────────────────────────────
def load_file(uploaded_file):
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb["IN-OUT"]
    headers = [cell.value for cell in ws[5]]
    data = [list(row) for row in ws.iter_rows(min_row=6, values_only=True)]
    df = pd.DataFrame(data, columns=headers)
    df.columns = [norm(c) if c else "" for c in df.columns]
    return df


def resolve_col(df, candidates):
    for c in candidates:
        if c in df.columns:
            return c
    return None


def get_cols(df):
    return {
        "carrier":       resolve_col(df, ["carrier"]),
        "xdock":         resolve_col(df, ["xdock"]),
        "folio":         resolve_col(df, ["folio almacen origen"]),
        "estatus":       resolve_col(df, ["estatus"]),
        "est_salida":    resolve_col(df, ["estatus salida", "estatus de salida", "status salida"]),
        "fecha_salida":  resolve_col(df, ["fecha de salida"]),
        "id_sitio":      resolve_col(df, ["id sitio"]),
        "nombre_sitio":  resolve_col(df, ["nombre de sitio"]),
        "no_pallet":     resolve_col(df, ["no. de pallet", "no de pallet"]),
        "tipo_pallet":   resolve_col(df, ["tipo de pallet"]),
        "tipo_material": resolve_col(df, ["tipo de material"]),
        "desc_material": resolve_col(df, ["descripcion material"]),
    }


# ─────────────────────────────────────────────────────────────────────────────
#  PIPELINE STEPS
# ─────────────────────────────────────────────────────────────────────────────
def filter_salidas(df, cols):
    mask = (
        df[cols["estatus"]].apply(lambda x: norm(x) == "salida")
        | df[cols["est_salida"]].apply(lambda x: norm(x) == "salida")
        | df[cols["fecha_salida"]].notna()
    )
    return df[~mask].copy(), df[mask].copy()


def normalize_xdock_name(val):
    return XDOCK_ALIASES.get(norm(val), val)


def filter_ens(df, cols):
    # Eliminate Selecciona... rows immediately — never infer, always drop
    mask_selecciona = df[cols["xdock"]].apply(lambda x: is_selecciona(x))
    df = df[~mask_selecciona].copy()

    df[cols["xdock"]] = df[cols["xdock"]].apply(
        lambda x: normalize_xdock_name(x) if not is_empty_xdock(x) else x
    )
    mask_ens   = df[cols["xdock"]].apply(lambda x: contains_ens(x))
    mask_empty = df[cols["xdock"]].apply(lambda x: is_empty_xdock(x))
    keep = mask_ens | mask_empty
    return df[keep].copy(), df[~keep].copy()


def infer_xdock_from_folio(folio):
    if not folio:
        return None
    f = str(folio).strip().upper()
    for prefix, xd in FOLIO_XDOCK:
        if f.startswith(prefix.upper()):
            return xd
    return None


def infer_xdock_from_id(id_sitio):
    if not id_sitio:
        return None
    s = str(id_sitio).strip().upper()
    if s == "HMTY0150":
        return "__DELETE__"
    if "JUMPER" in s.lower():
        return "__DELETE__"
    for prefix, xd in ID_SITIO_PREFIXES.items():
        if s.startswith(prefix.upper()):
            return xd
    return None


def infer_xdock_from_keywords(*fields):
    combined = " ".join(norm(str(f)) for f in fields if f)
    if "devolucion logistica inversa la paz" in combined:
        return "Gaso- La Paz-E-NS"
    for kw, xd in KEYWORD_XDOCK.items():
        if kw in combined:
            return xd
    return None


def fill_xdock(df, cols, saved_decisions: dict):
    """
    Fill empty XDOCK values.
    saved_decisions: {id_sitio_key -> {"action": "assign"|"delete", "xdock": "..."}}
    Returns (df_active, df_pending, n_auto_assigned, n_auto_deleted, n_decision_applied)
    """
    delete_idx  = []
    pending_idx = []
    n_auto  = 0
    n_saved = 0

    for idx, row in df.iterrows():
        xd = row[cols["xdock"]]
        if not is_empty_xdock(xd):
            continue

        folio     = row[cols["folio"]]
        id_sitio  = row[cols["id_sitio"]]
        nom_sitio = row[cols["nombre_sitio"]]
        desc      = row.get(cols["desc_material"], "")
        key       = str(id_sitio).strip().upper() if id_sitio else "__NOKEY__"

        # ── Check saved decision first ────────────────────────────────────
        if key in saved_decisions:
            dec = saved_decisions[key]
            if dec["action"] == "delete":
                delete_idx.append(idx)
                n_saved += 1
                continue
            elif dec["action"] == "assign" and dec.get("xdock"):
                df.at[idx, cols["xdock"]] = dec["xdock"]
                n_saved += 1
                continue

        # ── Auto-inference ────────────────────────────────────────────────
        result = infer_xdock_from_folio(folio)
        if result is None:
            result = infer_xdock_from_id(id_sitio)
        if result is None:
            result = infer_xdock_from_keywords(id_sitio, nom_sitio, folio, desc)

        if result == "__DELETE__":
            delete_idx.append(idx)
            n_auto += 1
        elif result:
            df.at[idx, cols["xdock"]] = result
            n_auto += 1
        else:
            pending_idx.append(idx)

    df = df.drop(index=delete_idx)
    df_pending = df.loc[df.index.isin(pending_idx)].copy()
    df_active  = df[~df.index.isin(pending_idx)].copy()
    return df_active, df_pending, n_auto, n_saved


def final_ens_filter(df, cols):
    mask = df[cols["xdock"]].apply(lambda x: contains_ens(str(x)))
    return df[mask].copy(), df[~mask].copy()


def fix_pallets(df, cols):
    def calc(row):
        if str(row[cols["id_sitio"]]).strip().upper() == "JALZAP0981":
            return 14
        val = row[cols["no_pallet"]]
        if pd.isna(val) or val == "" or val is None:
            return 1
        try:
            return 1  # any non-zero value → 1
        except Exception:
            return 1
    df[cols["no_pallet"]] = df.apply(calc, axis=1)
    return df


# Reglas para inferir tipo de pallet desde descripcion/clasificacion cuando
# el campo tipo_pallet está vacío. Orden importa: más específico primero.
DESC_PALLET_RULES = [
    # SOBREDIMENSIONADA
    ("antena gd",          "SOBREDIMENSIONADA"),
    ("antena grande",      "SOBREDIMENSIONADA"),
    ("antenas grandes",    "SOBREDIMENSIONADA"),
    ("galvanizado",        "SOBREDIMENSIONADA"),
    ("soporte",            "SOBREDIMENSIONADA"),   # cubre soportes, soporte desmontaje, etc.
    ("obra civil",         "SOBREDIMENSIONADA"),
    ("civil work",         "SOBREDIMENSIONADA"),
    ("base metalica",      "SOBREDIMENSIONADA"),
    ("base metálica",      "SOBREDIMENSIONADA"),
    ("vallen",             "SOBREDIMENSIONADA"),
    ("outdoor",            "SOBREDIMENSIONADA"),
    # EUROPALLET
    ("dsv",                "EUROPALLET"),          # pallet dsv / 1 pallet dsv / dsv-vallen
    ("hw sitio completo",  "EUROPALLET"),
    ("hw",                 "EUROPALLET"),          # HW solo o con tabuladores
    ("sde",                "EUROPALLET"),
    # ESTANDAR
    ("antena md",          "ESTANDAR"),
    ("antena mediana",     "ESTANDAR"),
    ("antena ch",          "EUROPALLET"),
    ("antena chica",       "EUROPALLET"),
    ("antena",             "ESTANDAR"),            # antena/antenas genérico → ESTANDAR
    ("gabinete",           "ESTANDAR"),
    ("complemento",        "ESTANDAR"),
    ("refurbish",          "ESTANDAR"),
    ("refaccion",          "ESTANDAR"),
    ("miscelaneo",         "ESTANDAR"),
    ("varios",             "ESTANDAR"),
    ("scrap",              "ESTANDAR"),
    ("logistica",          "ESTANDAR"),
    ("caja",               "ESTANDAR"),
    ("herraje",            "ESTANDAR"),
    ("sobrante",           "ESTANDAR"),
    ("cintillo",           "ESTANDAR"),
    ("bateria",            "ESTANDAR"),
    ("cable",              "ESTANDAR"),
    ("indoor",             "ESTANDAR"),
    ("rru",                "ESTANDAR"),
]


# Jerarquía de tamaño para elegir el tipo más grande cuando hay múltiples matches
TIPO_RANK = {"ESTANDAR": 1, "EUROPALLET": 2, "SOBREDIMENSIONADA": 3}


def _infer_tipo_from_desc(*fields):
    """
    Evalúa TODOS los segmentos del texto (separados por tab, coma, pipe, etc.)
    contra DESC_PALLET_RULES y devuelve el tipo MÁS GRANDE que encuentre.
    Si no hay ninguna coincidencia → "ESTANDAR" por default.
    """
    combined = " | ".join(norm(str(f)) for f in fields if f and str(f) not in ("nan", "none", ""))
    if not combined.strip():
        return "ESTANDAR"

    # Split on common separators so each segmento se evalúa por separado
    import re as _re
    segments = _re.split(r"[	,|/\n]+", combined)

    best = "ESTANDAR"
    for seg in segments:
        seg = seg.strip()
        if not seg:
            continue
        for keyword, tipo in DESC_PALLET_RULES:
            if keyword in seg:
                if TIPO_RANK.get(tipo, 0) > TIPO_RANK.get(best, 0):
                    best = tipo
                break  # una regla por segmento es suficiente
    return best


def normalize_tipo_pallet(df, cols):
    """
    1. Si tipo_pallet ya tiene valor reconocido → normalizar con PALLET_NORM.
    2. Si está vacío → inferir desde descripcion_material y clasificacion,
       tomando el tipo MÁS GRANDE entre todos los segmentos.
    3. Si nada coincide → ESTANDAR por default (nunca queda vacío).
    """
    desc_col  = cols.get("desc_material")
    clasi_col = None
    for c in df.columns:
        if "clasif" in c:
            clasi_col = c
            break

    resultado_inferencia = []

    def fix(row):
        val = row[cols["tipo_pallet"]]
        is_empty = pd.isna(val) or str(val).strip() in ("", "None", "nan")

        if not is_empty:
            n = norm(val)
            for k, v in PALLET_NORM.items():
                if k in n:
                    return v
            return "ESTANDAR"

        # Empty → infer (always returns a value, never None)
        desc  = row[desc_col]  if desc_col  and desc_col  in row.index else None
        clasi = row[clasi_col] if clasi_col and clasi_col in row.index else None
        inferred = _infer_tipo_from_desc(desc, clasi)

        resultado_inferencia.append({
            "ID_SITIO":      row.get(cols["id_sitio"], ""),
            "DESC":          str(desc)[:60],
            "CLASIF":        str(clasi)[:40],
            "TIPO_INFERIDO": inferred,
        })
        return inferred

    df[cols["tipo_pallet"]] = df.apply(fix, axis=1)

    global _last_tipo_inferences
    _last_tipo_inferences = resultado_inferencia

    return df


_last_tipo_inferences = []   # populated by normalize_tipo_pallet


def calc_m2(df, cols):
    def m2(row):
        tp  = str(row[cols["tipo_pallet"]]).upper().strip()
        np_ = float(row[cols["no_pallet"]]) if row[cols["no_pallet"]] else 0
        return round(np_ * M2_TIPO.get(tp, 1.44) * 1.20, 4)
    df["M2"] = df.apply(m2, axis=1)
    return df


def assign_region(df, cols):
    df["REGION"] = df[cols["xdock"]].map(REGION_MAP).fillna("SIN REGIÓN")
    df["CIUDAD"] = df[cols["xdock"]].map(CIUDAD_MAP).fillna("")
    return df


# ─────────────────────────────────────────────────────────────────────────────
#  MASTER PIPELINE
# ─────────────────────────────────────────────────────────────────────────────
def run_pipeline(uploaded_file, saved_decisions: dict):
    logs = []
    df_raw = load_file(uploaded_file)
    # Keep raw copy (with SALIDAS) for crossdock deep analysis
    df_raw_full = df_raw.copy()
    # Re-map normalized col names to original-style names for crossdock analysis
    col_remap = {v: k for k, v in {
        "CARRIER": "CARRIER", "XDOCK": "XDOCK",
        "FECHA DE INGRESO": "FECHA DE INGRESO", "HORA DE INGRESO": "HORA DE INGRESO",
        "FOLIO ALMACEN ORIGEN": "FOLIO ALMACEN ORIGEN", "ESTATUS": "ESTATUS",
        "CLASIFICACION DE MATERIAL": "CLASIFICACION DE MATERIAL",
        "TIPO DE MATERIAL": "TIPO DE MATERIAL", "ID SITIO": "ID SITIO",
        "NOMBRE DE SITIO": "NOMBRE DE SITIO",
        "FOLIO CLIENTE ( PDM )": "FOLIO CLIENTE ( PDM )", "ID PALLET": "ID PALLET",
        "NO. DE PALLET": "NO. DE PALLET", "TIPO DE PALLET": "TIPO DE PALLET",
        "FOLIO WMS GASO ( UNICO X PALLET )": "FOLIO WMS GASO ( UNICO X PALLET )",
        "DESCRIPCION MATERIAL": "DESCRIPCION MATERIAL", "PROYECTO": "PROYECTO",
        "SUB-PROYECTO": "SUB-PROYECTO", "FOLIO ENRUTADO CLIENTE": "FOLIO ENRUTADO CLIENTE",
        "ESTATUS SALIDA": "ESTATUS SALIDA", "DIAS INV.": "DIAS INV.",
        "FECHA DE SALIDA": "FECHA DE SALIDA", "HORA SALIDA": "HORA SALIDA",
        "NOMBRE ASP": "NOMBRE ASP", "NOMBRE OPERADOR": "NOMBRE OPERADOR",
        "PLACAS": "PLACAS", "OBSERVACIONES": "OBSERVACIONES",
        "PALLETS SALIDA": "PALLETS SALIDA", "EXISTENCIA REAL": "EXISTENCIA REAL",
        "DIAS QUE DURO EN INVENTARIO": "DIAS QUE DURO EN INVENTARIO",
    }.items()}
    # Restore original column names (un-normalize) so crossdock functions can find them
    from unidecode import unidecode as _ud
    def _unnorm(col):
        for orig in ["CARRIER","XDOCK","FECHA DE INGRESO","HORA DE INGRESO",
                     "FOLIO ALMACEN ORIGEN","ESTATUS","CLASIFICACION DE MATERIAL",
                     "TIPO DE MATERIAL","ID SITIO","NOMBRE DE SITIO",
                     "FOLIO CLIENTE ( PDM )","ID PALLET","NO. DE PALLET",
                     "TIPO DE PALLET","FOLIO WMS GASO ( UNICO X PALLET )",
                     "DESCRIPCION MATERIAL","PROYECTO","SUB-PROYECTO",
                     "FOLIO ENRUTADO CLIENTE","ESTATUS SALIDA","DIAS INV.",
                     "FECHA DE SALIDA","HORA SALIDA","NOMBRE ASP","NOMBRE OPERADOR",
                     "PLACAS","OBSERVACIONES","PALLETS SALIDA","EXISTENCIA REAL",
                     "DIAS QUE DURO EN INVENTARIO"]:
            if _ud(orig).strip().lower() == col.strip().lower():
                return orig
        return col
    df_raw_full.columns = [_unnorm(c) for c in df_raw_full.columns]
    logs.append(f"✅ Archivo cargado: {len(df_raw):,} registros totales")

    cols = get_cols(df_raw)

    df, rem_sal = filter_salidas(df_raw, cols)
    logs.append(f"🗑️  Salidas eliminadas: {len(rem_sal):,} | Restantes: {len(df):,}")

    df, rem_ens = filter_ens(df, cols)
    logs.append(f"🗑️  Sin E-NS eliminados: {len(rem_ens):,} | Restantes: {len(df):,}")

    df, df_pending, n_auto, n_saved = fill_xdock(df, cols, saved_decisions)
    logs.append(f"🔍 XDOCK resueltos automáticamente: {n_auto} | Por decisión guardada: {n_saved} | Pendientes nuevos: {len(df_pending)}")

    df, rem_fin = final_ens_filter(df, cols)
    logs.append(f"🗑️  Filtro final E-NS: {len(rem_fin)} eliminados | Activos: {len(df):,}")

    # Separar consolidados (NO. PALLET = 0) ANTES de fix_pallets
    # Son material pequeno encima de otra tarima: no impacta capacidad
    def _is_zero(val):
        try:
            return float(val) == 0
        except (TypeError, ValueError):
            return False

    mask_consol = df[cols["no_pallet"]].apply(_is_zero)
    df_consol   = df[mask_consol].copy()
    df          = df[~mask_consol].copy()
    logs.append(f"📌 Sitios consolidados (pallet=0): {len(df_consol)} | Para capacidad: {len(df):,}")

    df = fix_pallets(df, cols)
    df = normalize_tipo_pallet(df, cols)

    # Separar filas cuyo tipo_pallet no pudo inferirse → revisión manual de tipo
    n_inferred = len(_last_tipo_inferences)
    if n_inferred:
        logs.append(f"🔎 Tipo Pallet inferido por descripción: {n_inferred} filas")

    df = calc_m2(df, cols)
    df = assign_region(df, cols)

    # Rellenar tipo_material vacío con SIN CLASIFICAR
    mat_col = cols["tipo_material"]
    df[mat_col] = df[mat_col].fillna("SIN CLASIFICAR")
    df[mat_col] = df[mat_col].apply(
        lambda x: "SIN CLASIFICAR" if str(x).strip() in ("", "None", "nan") else x
    )

    # Enriquecer consolidados (sin M2 ni conteo de pallet)
    df_consol = normalize_tipo_pallet(df_consol, cols)
    df_consol["REGION"] = df_consol[cols["xdock"]].map(REGION_MAP).fillna("SIN REGION")
    df_consol["CIUDAD"] = df_consol[cols["xdock"]].map(CIUDAD_MAP).fillna("")
    df_consol["M2"]     = 0.0
    df["CONSOLIDADO"]       = False
    df_consol["CONSOLIDADO"] = True

    logs.append(f"📦 Pallets activos: {int(df[cols['no_pallet']].sum()):,} | M²: {df['M2'].sum():,.2f} | Consolidados: {len(df_consol)}")

    return df, df_consol, df_pending, logs, cols


# ─────────────────────────────────────────────────────────────────────────────
#  REPORT BUILDERS
# ─────────────────────────────────────────────────────────────────────────────
def build_client_report(df, cols):
    grp = df.groupby([cols["xdock"], cols["carrier"], cols["tipo_material"]]).agg(
        NO_PALLETS=(cols["no_pallet"], "sum"),
        M2=("M2", "sum"),
    ).reset_index()
    grp["CIUDAD"]    = grp[cols["xdock"]].map(CIUDAD_MAP).fillna(grp[cols["xdock"]])
    grp["REGION"]    = grp[cols["xdock"]].map(REGION_MAP).fillna("SIN REGIÓN")
    grp["CAPACIDAD"] = grp[cols["xdock"]].map(CAPACIDADES).fillna(0)
    grp["PCT_OCP"]   = grp.apply(
        lambda r: round(r["M2"] / r["CAPACIDAD"], 4) if r["CAPACIDAD"] > 0 else 0, axis=1
    )
    return grp



# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL EXPORT  (4 sheets: IN-OUT LIMPIO · REPORTE CLIENTE · RESUMEN EJECUTIVO · SITIOS CONSOLIDADOS)
# ─────────────────────────────────────────────────────────────────────────────
HEX_BLUE  = "1A3A6B"
HEX_LIGHT = "2E6DB4"
HEX_WHITE = "FFFFFF"
HEX_LGRAY = "EBF0F7"
HEX_GREEN = "1E8449"
HEX_AMBER = "E67E22"
HEX_RED   = "C0392B"
HEX_DRED  = "7B241C"

_thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

HDR_FONT   = Font(name="Calibri", bold=True, color=HEX_WHITE, size=10)
HDR_FILL   = PatternFill("solid", fgColor=HEX_BLUE)
DATA_FONT  = Font(name="Calibri", size=9)
CTR        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT        = Alignment(horizontal="left",   vertical="center")


def _pct_fill(pct):
    if pct > 1.00: return PatternFill("solid", fgColor=HEX_DRED)
    if pct > 0.90: return PatternFill("solid", fgColor=HEX_RED)
    if pct > 0.70: return PatternFill("solid", fgColor=HEX_AMBER)
    return PatternFill("solid", fgColor=HEX_GREEN)


def _alt_fill(i):
    return PatternFill("solid", fgColor=HEX_LGRAY if i % 2 else HEX_WHITE)


def _title_block(ws, title, subtitle, fecha, n_cols=18):
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 16
    end = get_column_letter(n_cols)
    for row, val, size, fg, bg in [
        (1, "GASO COMUNICACIONES",  15, HEX_WHITE, HEX_BLUE),
        (2, title,                  11, HEX_WHITE, HEX_LIGHT),
        (3, f"{subtitle}   |   Fecha generación: {fecha}", 9, "555555", "F4F6F9"),
    ]:
        ws.merge_cells(f"A{row}:{end}{row}")
        c = ws[f"A{row}"]
        c.value     = val
        c.font      = Font(name="Calibri", bold=(row < 3), size=size, color=fg)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = CTR


def _write_table(ws, df, start_row, freeze=True):
    cols = list(df.columns)
    for ci, h in enumerate(cols, 1):
        c = ws.cell(row=start_row, column=ci, value=str(h))
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.border = BORDER; c.alignment = CTR
    for ri, (_, row) in enumerate(df.iterrows()):
        er = start_row + 1 + ri
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=er, column=ci, value=val)
            c.font = DATA_FONT; c.border = BORDER
            c.fill = _alt_fill(ri); c.alignment = LFT
    # auto-width
    for ci, col in enumerate(cols, 1):
        try:
            mx = max(len(str(col)),
                     df.iloc[:, ci-1].astype(str).str.len().max() if len(df) else 10)
        except Exception:
            mx = 14
        ws.column_dimensions[get_column_letter(ci)].width = min(mx + 3, 42)
    if freeze:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)



def build_original_format_excel(df_clean, df_consol, cols):
    """Reproduce the original IN-OUT workbook header style with the clean data."""
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = "IN-OUT"
    fecha = datetime.date.today().strftime("%d/%m/%Y")

    # ── Colour palette matching the original ──────────────────────────────────
    C_DARK_BLUE = "002060"   # Row 1 / headers INGRESO
    C_DARK_GRN  = "0B6A0B"   # headers SALIDA
    C_GOLD      = "FFD966"   # headers INVENTARIO
    C_LT_BLUE   = "D9E1F2"   # sub-header INGRESO
    C_LT_GRN    = "C6EFCE"   # sub-header SALIDA
    C_LT_GOLD   = "FFF2CC"   # sub-header INVENTARIO
    C_WHITE     = "FFFFFF"

    def _mk_fill(hex_col):
        return PatternFill("solid", fgColor=hex_col)

    def _mk_font(bold=False, color=C_WHITE, size=11):
        return Font(name="Calibri", bold=bold, color=color, size=size)

    def _ctr():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _lft():
        return Alignment(horizontal="left", vertical="center")

    # Column layout: map to original widths (A=CARRIER … )
    # We output only the active/relevant columns
    out_cols_names = [
        "CARRIER", "XDOCK", "FECHA DE INGRESO", "HORA DE INGRESO",
        "FOLIO ALMACEN ORIGEN", "ESTATUS", "CLASIFICACION DE MATERIAL",
        "TIPO DE MATERIAL", "ID SITIO", "NOMBRE DE SITIO",
        "FOLIO CLIENTE ( PDM )", "ID PALLET", "NO. DE PALLET",
        "TIPO DE PALLET", "FOLIO WMS GASO ( UNICO X PALLET )",
        "DESCRIPCION MATERIAL", "PROYECTO", "SUB-PROYECTO",
        "FOLIO ENRUTADO CLIENTE", "ESTATUS SALIDA",
        "REGION", "CIUDAD", "M2", "CONSOLIDADO",
    ]
    # Map to df columns (normalized)
    col_lookup = {
        "CARRIER":                     cols["carrier"],
        "XDOCK":                       cols["xdock"],
        "FECHA DE INGRESO":            resolve_col(df_clean, ["fecha de ingreso"]),
        "HORA DE INGRESO":             resolve_col(df_clean, ["hora de ingreso"]),
        "FOLIO ALMACEN ORIGEN":        cols["folio"],
        "ESTATUS":                     cols["estatus"],
        "CLASIFICACION DE MATERIAL":   resolve_col(df_clean, ["clasificacion de material"]),
        "TIPO DE MATERIAL":            cols["tipo_material"],
        "ID SITIO":                    cols["id_sitio"],
        "NOMBRE DE SITIO":             cols["nombre_sitio"],
        "FOLIO CLIENTE ( PDM )":       resolve_col(df_clean, ["folio cliente ( pdm )"]),
        "ID PALLET":                   resolve_col(df_clean, ["id pallet"]),
        "NO. DE PALLET":               cols["no_pallet"],
        "TIPO DE PALLET":              cols["tipo_pallet"],
        "FOLIO WMS GASO ( UNICO X PALLET )": resolve_col(df_clean, ["folio wms gaso ( unico x pallet )"]),
        "DESCRIPCION MATERIAL":        cols.get("desc_material"),
        "PROYECTO":                    resolve_col(df_clean, ["proyecto"]),
        "SUB-PROYECTO":                resolve_col(df_clean, ["sub-proyecto"]),
        "FOLIO ENRUTADO CLIENTE":      resolve_col(df_clean, ["folio enrutado cliente"]),
        "ESTATUS SALIDA":              cols["est_salida"],
        "REGION":                      "REGION",
        "CIUDAD":                      "CIUDAD",
        "M2":                          "M2",
        "CONSOLIDADO":                 "CONSOLIDADO",
    }
    # Original column widths (approx)
    orig_widths = {
        "CARRIER": 16.8, "XDOCK": 14.8, "FECHA DE INGRESO": 16.8,
        "HORA DE INGRESO": 12.8, "FOLIO ALMACEN ORIGEN": 22.8,
        "ESTATUS": 14.8, "CLASIFICACION DE MATERIAL": 26.8,
        "TIPO DE MATERIAL": 20.8, "ID SITIO": 18.8, "NOMBRE DE SITIO": 28.8,
        "FOLIO CLIENTE ( PDM )": 20.8, "ID PALLET": 16.8,
        "NO. DE PALLET": 16.8, "TIPO DE PALLET": 16.8,
        "FOLIO WMS GASO ( UNICO X PALLET )": 28.8,
        "DESCRIPCION MATERIAL": 32.8, "PROYECTO": 18.8, "SUB-PROYECTO": 18.8,
        "FOLIO ENRUTADO CLIENTE": 22.8, "ESTATUS SALIDA": 16.8,
        "REGION": 16.8, "CIUDAD": 14.8, "M2": 10.0, "CONSOLIDADO": 12.0,
    }

    n_cols = len(out_cols_names)
    last_col_letter = get_column_letter(n_cols)

    # Row 1: REPORTE IN-OUT LIMPIO (dark blue, merged)
    ws.merge_cells(f"A1:{last_col_letter}1")
    c = ws["A1"]
    c.value = "REPORTE IN-OUT"
    c.font  = _mk_font(bold=True, color=C_WHITE, size=13)
    c.fill  = _mk_fill(C_DARK_BLUE)
    c.alignment = _ctr()
    ws.row_dimensions[1].height = 22

    # Row 2: GASO COMUNICACIONES
    ws.merge_cells(f"A2:{last_col_letter}2")
    c = ws["A2"]
    c.value = f"GASO COMUNICACIONES  –  Base Limpia Procesada  |  {fecha}"
    c.font  = _mk_font(bold=True, color=C_DARK_BLUE, size=11)
    c.fill  = _mk_fill("F4F6F9")
    c.alignment = _ctr()
    ws.row_dimensions[2].height = 18

    # Row 3: Section labels INGRESO / SALIDA / INVENTARIO
    # INGRESO spans cols 1-20, SALIDA none (no output), REGION cols 21+
    ingreso_end   = min(20, n_cols)
    region_start  = 21
    ws.merge_cells(f"A3:{get_column_letter(ingreso_end)}3")
    c = ws["A3"]; c.value = "INGRESO"
    c.font = _mk_font(bold=True); c.fill = _mk_fill(C_DARK_BLUE); c.alignment = _ctr()
    if region_start <= n_cols:
        ws.merge_cells(f"{get_column_letter(region_start)}3:{last_col_letter}3")
        c = ws[f"{get_column_letter(region_start)}3"]; c.value = "INVENTARIO / MÉTRICAS"
        c.font = _mk_font(bold=True, color="333333"); c.fill = _mk_fill(C_GOLD); c.alignment = _ctr()
    ws.row_dimensions[3].height = 18

    # Row 4: Sub-headers INPUT / INVENTARIO
    ws.merge_cells(f"A4:{get_column_letter(ingreso_end)}4")
    c = ws["A4"]; c.value = "INPUT"
    c.font = _mk_font(bold=True, color=C_DARK_BLUE); c.fill = _mk_fill(C_LT_BLUE); c.alignment = _ctr()
    if region_start <= n_cols:
        ws.merge_cells(f"{get_column_letter(region_start)}4:{last_col_letter}4")
        c = ws[f"{get_column_letter(region_start)}4"]; c.value = "INVENTARIO"
        c.font = _mk_font(bold=True, color="333333"); c.fill = _mk_fill(C_LT_GOLD); c.alignment = _ctr()
    ws.row_dimensions[4].height = 18

    # Row 5: Column headers
    ws.row_dimensions[5].height = 30
    for ci, col_name in enumerate(out_cols_names, 1):
        is_region_col = ci >= region_start
        c = ws.cell(row=5, column=ci, value=col_name)
        c.font  = _mk_font(bold=True, color=C_WHITE, size=11)
        c.fill  = _mk_fill(C_GOLD if is_region_col else C_DARK_BLUE)
        if is_region_col:
            c.font = _mk_font(bold=True, color="333333", size=11)
        c.alignment = _ctr()
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = orig_widths.get(col_name, 14.0)

    # Data rows - combine clean + consolidados
    df_all = pd.concat([df_clean, df_consol], ignore_index=True)
    thin = Side(style="thin", color="CCCCCC")
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ri, (_, row) in enumerate(df_all.iterrows()):
        er = 6 + ri
        alt = ri % 2 == 1
        base_fill = PatternFill("solid", fgColor="EBF0F7" if alt else "FFFFFF")
        for ci, col_name in enumerate(out_cols_names, 1):
            df_col = col_lookup.get(col_name)
            val = row[df_col] if df_col and df_col in row.index else None
            c = ws.cell(row=er, column=ci, value=val)
            c.font      = Font(name="Calibri", size=10)
            c.alignment = _lft()
            c.border    = data_border
            c.fill      = base_fill

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:{last_col_letter}5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def build_excel_output(df_clean, df_consol, cols, region_filter="Todas"):
    wb   = openpyxl.Workbook()
    fecha = datetime.date.today().strftime("%d/%m/%Y")

    xdock_col  = cols["xdock"]
    carrier_col = cols["carrier"]
    mat_col    = cols["tipo_material"]
    pallet_col = cols["no_pallet"]

    # Apply region filter
    if region_filter != "Todas":
        df_rep    = df_clean[df_clean["REGION"] == region_filter].copy()
        df_c_rep  = df_consol[df_consol[xdock_col].isin(
                        [xd for xd in CAPACIDADES if REGION_MAP.get(xd) == region_filter]
                    )].copy() if len(df_consol) > 0 else df_consol.copy()
        xdocks_rep = [xd for xd in sorted(CAPACIDADES.keys()) if REGION_MAP.get(xd) == region_filter]
        region_label = region_filter
    else:
        df_rep     = df_clean.copy()
        df_c_rep   = df_consol.copy()
        xdocks_rep = sorted(CAPACIDADES.keys())
        region_label = "Todas las Regiones"

    # ── 1. IN-OUT LIMPIO ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "IN-OUT LIMPIO"
    _title_block(ws1, "BASE DE DATOS – IN-OUT LIMPIO",
                 f"Inventario activo depurado y enriquecido  |  Región: {region_label}",
                 fecha, n_cols=min(30, len(df_rep.columns)))
    _write_table(ws1, df_rep.reset_index(drop=True), start_row=5)

    # ── 2. REPORTE CLIENTE ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("REPORTE CLIENTE")
    ws2.sheet_view.showGridLines = False

    carriers = sorted(df_rep[carrier_col].dropna().unique())
    tipos    = sorted(df_rep[mat_col].dropna().unique())
    xdocks   = xdocks_rep

    # Build header columns
    exec_cols  = ["CIUDAD", "REGIÓN", "CAPACIDAD M²"]
    for car in carriers:
        for tp in tipos:
            exec_cols.append(f"{car}\n{tp[:18]}\nPallets")
        exec_cols.append(f"{car}\nTotal Pallets")
        exec_cols.append(f"{car}\nM² Ocupados")
    exec_cols += ["TOTAL\nPALLETS", "TOTAL M²\nOCUPADOS", "% OCUPACIÓN", "DISPONIBLE M²", "STATUS"]
    for car in carriers:
        exec_cols.append(f"% M²\n{car}\ndel Total")

    HDR_ROW = 5
    _title_block(ws2, "REPORTE SEMANAL DE OCUPACIÓN – CLIENTE",
                 f"Resumen ejecutivo por XDOCK · Carrier · Tipo de Material  |  Región: {region_label}", fecha, n_cols=len(exec_cols))

    ws2.row_dimensions[HDR_ROW].height = 40
    for ci, h in enumerate(exec_cols, 1):
        c = ws2.cell(row=HDR_ROW, column=ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.border = BORDER; c.alignment = CTR
        ws2.column_dimensions[get_column_letter(ci)].width = 13
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 12
    # Wider columns for % carrier cols and set a distinguishing fill
    for car_i in range(len(carriers)):
        car_ci = len(exec_cols) - len(carriers) + car_i + 1
        ws2.column_dimensions[get_column_letter(car_ci)].width = 14
        c_h = ws2.cell(row=HDR_ROW, column=car_ci)
        c_h.fill = PatternFill("solid", fgColor="2E6DB4")   # accent blue
        c_h.font = Font(name="Calibri", bold=True, color=HEX_WHITE, size=9)

    total_pal_all = 0
    total_m2_all  = 0.0

    for ri, xd in enumerate(xdocks):
        er  = HDR_ROW + 1 + ri
        ciudad = CIUDAD_MAP.get(xd, xd)
        region = REGION_MAP.get(xd, "")
        cap    = CAPACIDADES.get(xd, 0)

        row_vals = [ciudad, region, cap]
        tot_pal = 0; tot_m2 = 0.0

        for car in carriers:
            sub = df_rep[(df_rep[xdock_col] == xd) & (df_rep[carrier_col] == car)]
            for tp in tipos:
                sub2 = sub[sub[mat_col] == tp]
                row_vals.append(int(sub2[pallet_col].sum()))
            c_pal = int(sub[pallet_col].sum())
            c_m2  = round(sub["M2"].sum(), 2)
            row_vals += [c_pal, c_m2]
            tot_pal += c_pal; tot_m2 += c_m2

        tot_m2 = round(tot_m2, 2)
        pct    = round(tot_m2 / cap, 4) if cap > 0 else 0
        disp   = round(cap - tot_m2, 2)
        status = ("SATURADO" if pct > 1.0 else "CRÍTICO" if pct > 0.90
                  else "ALERTA" if pct > 0.70 else "NORMAL")
        row_vals += [tot_pal, tot_m2, pct, disp, status]
        # % M² por carrier vs total M² ocupados del crossdock (AT&T + Telcel = 100%)
        # Solo supera 100% si el total de M² excede la capacidad del crossdock
        for car in carriers:
            car_m2 = df_rep[(df_rep[xdock_col] == xd) & (df_rep[carrier_col] == car)]["M2"].sum()
            row_vals.append(round(car_m2 / tot_m2, 4) if tot_m2 > 0 else 0)

        total_pal_all += tot_pal
        total_m2_all  += tot_m2

        n_fixed_tail = 5   # TOTAL PALLETS · TOTAL M2 · %OCP · DISPONIBLE · STATUS
        pct_col_idx  = len(exec_cols) - 2   # 0-based distance from end

        for ci, val in enumerate(row_vals, 1):
            c = ws2.cell(row=er, column=ci, value=val)
            c.border = BORDER; c.font = DATA_FONT
            c.fill   = _alt_fill(ri); c.alignment = CTR

        # Color % ocupación total
        n_cols_total = len(exec_cols)
        pct_ci = n_cols_total - 2 - len(carriers)   # before the carrier % cols
        c_pct  = ws2.cell(row=er, column=pct_ci)
        c_pct.number_format = "0.00%"
        c_pct.fill = _pct_fill(pct)
        c_pct.font = Font(name="Calibri", bold=True, color=HEX_WHITE, size=9)
        c_pct.alignment = CTR
        # Color % M² por carrier (light blue header fill, value formatted as %)
        for car_i in range(len(carriers)):
            car_pct_ci = n_cols_total - len(carriers) + car_i + 1
            c_cp = ws2.cell(row=er, column=car_pct_ci)
            c_cp.number_format = "0.00%"
            c_cp.font = Font(name="Calibri", bold=True, size=9,
                             color=HEX_WHITE if float(c_cp.value or 0) > 0 else "AAAAAA")
            c_cp.fill = PatternFill("solid", fgColor="2E6DB4")  # GASO_LIGHT blue
            c_cp.alignment = CTR

    # Totals row
    tr = HDR_ROW + 1 + len(xdocks_rep)
    total_cap = sum(CAPACIDADES[xd] for xd in xdocks_rep)
    pct_global = round(total_m2_all / total_cap, 4)

    ws2.cell(row=tr, column=1, value="TOTAL GLOBAL").font  = Font(name="Calibri", bold=True, color=HEX_WHITE, size=10)
    ws2.cell(row=tr, column=1).fill  = PatternFill("solid", fgColor=HEX_BLUE)
    ws2.cell(row=tr, column=1).alignment = CTR
    ws2.cell(row=tr, column=3, value=total_cap).font  = Font(name="Calibri", bold=True, color=HEX_WHITE, size=9)
    ws2.cell(row=tr, column=3).fill = PatternFill("solid", fgColor=HEX_BLUE)
    ws2.cell(row=tr, column=3).alignment = CTR

    nc = len(exec_cols)
    for ci, val in [(nc-3, total_pal_all), (nc-2, round(total_m2_all, 2))]:
        c = ws2.cell(row=tr, column=ci, value=val)
        c.font = Font(name="Calibri", bold=True, color=HEX_WHITE, size=9)
        c.fill = PatternFill("solid", fgColor=HEX_BLUE); c.alignment = CTR

    c_pg = ws2.cell(row=tr, column=nc-1, value=pct_global)
    c_pg.number_format = "0.00%"
    c_pg.fill = _pct_fill(pct_global)
    c_pg.font = Font(name="Calibri", bold=True, color=HEX_WHITE, size=9)
    c_pg.alignment = CTR

    ws2.freeze_panes = ws2.cell(row=HDR_ROW + 1, column=1)

    # ── 3. RESUMEN EJECUTIVO ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("RESUMEN EJECUTIVO")
    ws4.sheet_view.showGridLines = False
    _title_block(ws4, "RESUMEN EJECUTIVO DE OCUPACIÓN",
                 "Indicadores clave de desempeño del inventario en tiempo real", fecha, n_cols=8)

    total_pal   = int(df_rep[pallet_col].sum())
    total_m2    = round(df_rep["M2"].sum(), 2)
    disp_global = round(total_cap - total_m2, 2)

    # KPI block
    kpis = [
        ("Total Pallets en Inventario",   f"{total_pal:,}",         "unidades"),
        ("Total M² Ocupados",             f"{total_m2:,.2f}",        "m²"),
        ("Capacidad Total del Sistema",   f"{total_cap:,}",          "m²"),
        ("% Ocupación Global",            f"{pct_global*100:.1f}%",  ""),
        ("M² Disponibles",                f"{disp_global:,.2f}",     "m²"),
        ("Crossdocks activos",            f"{df_rep[xdock_col].nunique()}", "xdocks"),
    ]
    for ci, h in enumerate(["INDICADOR", "VALOR", "UNIDAD"], 1):
        c = ws4.cell(row=5, column=ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER; c.alignment = CTR
    ws4.column_dimensions["A"].width = 32
    ws4.column_dimensions["B"].width = 18
    ws4.column_dimensions["C"].width = 12
    for ri, (kpi, val, unit) in enumerate(kpis, 6):
        f = _alt_fill(ri)
        for ci, v in enumerate([kpi, val, unit], 1):
            c = ws4.cell(row=ri, column=ci, value=v)
            c.fill = f; c.border = BORDER
            c.font = Font(name="Calibri", bold=(ci == 1), size=10)
            c.alignment = CTR if ci > 1 else LFT

    # Occupancy table
    ws4.cell(row=13, column=1, value="OCUPACIÓN POR CROSSDOCK").font = \
        Font(name="Calibri", bold=True, size=11, color=HEX_BLUE)
    ocp_hdr = ["CIUDAD", "REGIÓN", "CAP. M²", "PALLETS", "M² OCUPADOS", "% OCUPACIÓN", "DISPONIBLE M²", "STATUS"]
    for ci, h in enumerate(ocp_hdr, 1):
        c = ws4.cell(row=14, column=ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER; c.alignment = CTR
        ws4.column_dimensions[get_column_letter(ci)].width = 16

    for ri, xd in enumerate(xdocks_rep, 15):
        ciudad = CIUDAD_MAP.get(xd, xd)
        cap    = CAPACIDADES.get(xd, 0)
        m2_ocp = round(df_rep[df_rep[xdock_col] == xd]["M2"].sum(), 2)
        pal    = int(df_rep[df_rep[xdock_col] == xd][pallet_col].sum())
        pct    = round(m2_ocp / cap, 4) if cap > 0 else 0
        disp   = round(cap - m2_ocp, 2)
        region = REGION_MAP.get(xd, "")
        status = ("SATURADO" if pct > 1.0 else "CRÍTICO" if pct > 0.90
                  else "ALERTA" if pct > 0.70 else "NORMAL")
        row_v = [ciudad, region, cap, pal, m2_ocp, pct, disp, status]
        f = _alt_fill(ri)
        for ci, val in enumerate(row_v, 1):
            c = ws4.cell(row=ri, column=ci, value=val)
            c.fill = f; c.border = BORDER; c.font = DATA_FONT; c.alignment = CTR
        c_pct = ws4.cell(row=ri, column=6)
        c_pct.number_format = "0.00%"
        c_pct.fill = _pct_fill(pct)
        c_pct.font = Font(name="Calibri", bold=True, color=HEX_WHITE, size=9)


    # ── 5. SITIOS CONSOLIDADOS ───────────────────────────────────────────────
    ws5 = wb.create_sheet("SITIOS CONSOLIDADOS")
    _title_block(ws5, "SITIOS CONSOLIDADOS",
                 "Material consolidado sobre otra tarima – no impacta en capacidad ni M²", fecha,
                 n_cols=min(20, len(df_consol.columns) if len(df_consol) > 0 else 8))
    if len(df_consol) > 0:
        # Drop helper column before export
        consol_export = df_consol.drop(columns=["CONSOLIDADO"], errors="ignore").reset_index(drop=True)
        _write_table(ws5, consol_export, start_row=5)
        # Summary by XDOCK
        sum_row = 5 + len(consol_export) + 3
        ws5.cell(row=sum_row, column=1, value="RESUMEN POR CROSSDOCK").font =             Font(name="Calibri", bold=True, size=10, color=HEX_BLUE)
        sum_hdr = ["CIUDAD", "XDOCK", "CARRIER", "SITIOS CONSOLIDADOS"]
        for ci, h in enumerate(sum_hdr, 1):
            c = ws5.cell(row=sum_row + 1, column=ci, value=h)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER; c.alignment = CTR
            ws5.column_dimensions[get_column_letter(ci)].width = 22
        grp_c = (df_c_rep.groupby([cols["xdock"], cols["carrier"]])
                 .size().reset_index(name="SITIOS"))
        for ri2, row2 in grp_c.iterrows():
            er2 = sum_row + 2 + ri2
            vals = [CIUDAD_MAP.get(row2[cols["xdock"]], row2[cols["xdock"]]),
                    row2[cols["xdock"]], row2[cols["carrier"]], row2["SITIOS"]]
            for ci, v in enumerate(vals, 1):
                c = ws5.cell(row=er2, column=ci, value=v)
                c.fill = _alt_fill(ri2); c.border = BORDER
                c.font = DATA_FONT; c.alignment = CTR
    else:
        ws5["A5"] = "No hay sitios consolidados en esta región."
        ws5["A5"].font = Font(name="Calibri", size=10, color="555555")


    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
#  PLOTLY CHARTS
# ─────────────────────────────────────────────────────────────────────────────

def _smart_analysis(xd, pct, total_pal, m2_ocp, cap, tipo_dist, carrier_dist=None):
    """
    Generate a contextual 2-3 sentence analysis for a crossdock.
    Varies by occupancy level, material mix, and carrier concentration.
    Never mentions internal calculation factors.
    """
    ciudad  = CIUDAD_MAP.get(xd, xd)
    pct100  = round(pct * 100, 1)
    disp    = round(cap - m2_ocp, 0)

    # ── Main sentence by occupancy level ────────────────────────────────────
    if pct > 1.20:
        main = (f"{ciudad} registra una saturación severa del {pct100}%, con {m2_ocp:.0f} m² "
                f"ocupados sobre una capacidad de {cap} m². Esta situación requiere revisión "
                f"inmediata del inventario: es probable que existan registros sin salida capturada "
                f"o material acumulado sin programa de despacho definido.")
    elif pct > 1.0:
        main = (f"{ciudad} supera su capacidad operativa con un {pct100}% de ocupación "
                f"({m2_ocp:.0f} m² vs {cap} m² disponibles). Se deben coordinar salidas "
                f"prioritarias y suspender nuevas recepciones hasta liberar espacio.")
    elif pct > 0.90:
        main = (f"{ciudad} opera en nivel crítico con {pct100}% de ocupación y solo "
                f"{disp:.0f} m² libres. Con este margen, cualquier entrada adicional "
                f"podría saturar el crossdock. Se recomienda programar salidas esta semana.")
    elif pct > 0.70:
        main = (f"{ciudad} se encuentra en zona de alerta con {pct100}% de ocupación "
                f"({m2_ocp:.0f}/{cap} m²). Tiene capacidad para absorber entregas menores, "
                f"pero se recomienda monitoreo diario y priorizar despachos pendientes.")
    elif pct > 0.40:
        main = (f"{ciudad} opera en niveles normales con {pct100}% de ocupación "
                f"({m2_ocp:.0f} m² ocupados, {disp:.0f} m² disponibles). "
                f"El crossdock tiene capacidad suficiente para recibir nuevas entregas sin restricción.")
    elif total_pal == 0:
        main = (f"{ciudad} no registra inventario activo en este corte. "
                f"Se recomienda verificar si hay entradas pendientes de captura "
                f"o si el crossdock se encuentra en periodo de baja operativa.")
    else:
        main = (f"{ciudad} tiene una ocupación baja del {pct100}% con {total_pal} pallets activos. "
                f"Cuenta con {disp:.0f} m² disponibles, lo que le da amplia holgura "
                f"para recibir material en las próximas semanas.")

    # ── Secondary insight: material mix ─────────────────────────────────────
    insight = ""
    if tipo_dist and total_pal > 0:
        sorted_tipos = sorted(tipo_dist.items(), key=lambda x: x[1], reverse=True)
        top_tipo, top_val = sorted_tipos[0]
        top_pct = round(top_val / total_pal * 100)
        if top_tipo == "SOBREDIMENSIONADA" and top_pct > 50:
            insight = (f" El {top_pct}% del inventario corresponde a material sobredimensionado, "
                       f"lo que explica el alto consumo de m² en relación al número de pallets.")
        elif top_tipo == "EUROPALLET" and top_pct > 50:
            insight = (f" Predomina material tipo Europallet ({top_pct}% del total), "
                       f"con una huella de espacio moderada por unidad.")
        elif top_tipo != "ESTANDAR" and top_pct > 60:
            insight = (f" El tipo de material predominante es {top_tipo} con {top_pct}% del inventario.")

    # ── Tertiary insight: carrier concentration ──────────────────────────────
    carrier_note = ""
    if carrier_dist and total_pal > 0:
        top_car  = max(carrier_dist, key=carrier_dist.get)
        top_c_pct = round(carrier_dist[top_car] / total_pal * 100)
        if top_c_pct >= 90:
            carrier_note = (f" El inventario es prácticamente exclusivo de {top_car} "
                            f"({top_c_pct}% de los pallets).")
        elif top_c_pct >= 70:
            carrier_note = f" {top_car} concentra el {top_c_pct}% del inventario en este crossdock."

    return main + insight + carrier_note


def generate_pdf(df_clean, df_consol, cols, region_filter="Todas"):
    """Build an executive PDF report with charts and smart analysis per crossdock."""
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage
    from reportlab.platypus import Table, TableStyle, PageBreak, HRFlowable, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
    from reportlab.lib.pagesizes import A4, landscape

    # Filter by region if needed
    if region_filter != "Todas":
        df_plot = df_clean[df_clean["REGION"] == region_filter].copy()
        xdocks_plot = [xd for xd in CAPACIDADES if REGION_MAP.get(xd) == region_filter]
    else:
        df_plot = df_clean.copy()
        xdocks_plot = list(CAPACIDADES.keys())

    xdock_col  = cols["xdock"]
    carrier_col = cols["carrier"]
    mat_col    = cols["tipo_material"]
    pallet_col = cols["no_pallet"]
    fecha_str  = datetime.date.today().strftime("%d de %B de %Y")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2.5*cm, bottomMargin=2*cm,
        title="Reporte Ejecutivo Gaso Comunicaciones",
        author="Gaso Comunicaciones",
    )

    # ── Styles ────────────────────────────────────────────────────────────────
    RL_BLUE  = colors.HexColor("#1A3A6B")
    RL_LIGHT = colors.HexColor("#2E6DB4")
    RL_ACCENT= colors.HexColor("#4A90D9")
    RL_GREEN = colors.HexColor("#1E8449")
    RL_AMBER = colors.HexColor("#E67E22")
    RL_RED   = colors.HexColor("#C0392B")
    RL_DRED  = colors.HexColor("#7B241C")
    RL_LGRAY = colors.HexColor("#F4F6F9")
    RL_WHITE = colors.white

    styles = getSampleStyleSheet()

    def _sty(name, **kw):
        return ParagraphStyle(name, **kw)

    S_COVER_TITLE = _sty("CoverTitle", fontSize=28, textColor=RL_WHITE,
                          fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=6)
    S_COVER_SUB   = _sty("CoverSub",   fontSize=13, textColor=colors.HexColor("#C8D8EC"),
                          fontName="Helvetica", alignment=TA_CENTER, spaceAfter=4)
    S_COVER_DATE  = _sty("CoverDate",  fontSize=10, textColor=colors.HexColor("#AABBCC"),
                          fontName="Helvetica", alignment=TA_CENTER)
    S_H1    = _sty("H1",   fontSize=14, textColor=RL_BLUE, fontName="Helvetica-Bold",
                   spaceBefore=14, spaceAfter=4, borderPad=0)
    S_H2    = _sty("H2",   fontSize=11, textColor=RL_LIGHT, fontName="Helvetica-Bold",
                   spaceBefore=8, spaceAfter=3)
    S_BODY  = _sty("Body", fontSize=9,  textColor=colors.HexColor("#333333"),
                   fontName="Helvetica", leading=14, alignment=TA_JUSTIFY, spaceAfter=6)
    S_SMALL = _sty("Small",fontSize=8,  textColor=colors.HexColor("#777777"),
                   fontName="Helvetica", leading=11)
    S_CTR   = _sty("Ctr",  fontSize=9,  textColor=colors.HexColor("#333333"),
                   fontName="Helvetica", alignment=TA_CENTER)
    S_BOLD  = _sty("Bold", fontSize=9,  textColor=RL_BLUE, fontName="Helvetica-Bold",
                   leading=13)

    def _hr():
        return HRFlowable(width="100%", thickness=1, color=RL_ACCENT, spaceAfter=6, spaceBefore=2)

    def _spacer(h=0.3):
        return Spacer(1, h*cm)

    def _fig_to_image(fig, width_cm=16, height_cm=8):
        """Export a plotly figure to a ReportLab Image flowable."""
        img_bytes = fig.to_image(format="png", width=int(width_cm/16*1200),
                                  height=int(height_cm/8*600), scale=2)
        return RLImage(io.BytesIO(img_bytes), width=width_cm*cm, height=height_cm*cm)

    story = []

    # Pre-compute KPIs used on cover and throughout
    region_label = region_filter if region_filter != "Todas" else "Todas las Regiones"
    total_pal  = int(df_plot[pallet_col].sum())
    total_m2   = round(df_plot["M2"].sum(), 2)
    cap_region = sum(CAPACIDADES[xd] for xd in xdocks_plot)
    pct_global = round(total_m2 / cap_region * 100, 1) if cap_region > 0 else 0
    disponible = round(cap_region - total_m2, 2)

    def _pct_color(p):
        if p > 100: return RL_DRED
        if p > 90:  return RL_RED
        if p > 70:  return RL_AMBER
        return RL_GREEN

    # ══════════════════════════════════════════════════════════════════════════
    # COVER PAGE  — logo above blue band, text inside blue band, KPIs below
    # ══════════════════════════════════════════════════════════════════════════

    # 1. Logo centered above the blue block
    if os.path.exists(LOGO_PATH):
        try:
            logo = RLImage(LOGO_PATH, width=3.8*cm, height=1.9*cm)
            logo.hAlign = "CENTER"
            story.append(_spacer(1.2))
            story.append(logo)
            story.append(_spacer(0.6))
        except Exception:
            story.append(_spacer(3))
    else:
        story.append(_spacer(4))

    # 2. Blue banner — each row is a separate cell so padding applies cleanly
    S_BT = _sty("BT", fontSize=26, textColor=RL_WHITE, fontName="Helvetica-Bold",
                alignment=TA_CENTER, leading=32, spaceBefore=0, spaceAfter=0)
    S_BS = _sty("BS", fontSize=12, textColor=colors.HexColor("#C8D8EC"),
                fontName="Helvetica", alignment=TA_CENTER, leading=18,
                spaceBefore=0, spaceAfter=0)
    S_BD = _sty("BD", fontSize=10, textColor=colors.HexColor("#AABBCC"),
                fontName="Helvetica", alignment=TA_CENTER, leading=14,
                spaceBefore=0, spaceAfter=0)

    banner_rows = [
        [Paragraph("GASO COMUNICACIONES", S_BT)],
        [Paragraph("Reporte Ejecutivo de Ocupación de Inventario", S_BS)],
        [Paragraph(f"Región: {region_label}", S_BS)],
        [Paragraph(fecha_str, S_BD)],
    ]
    banner_padding = [
        ("BACKGROUND",    (0,0), (-1,-1), RL_BLUE),
        ("LEFTPADDING",   (0,0), (-1,-1), 24),
        ("RIGHTPADDING",  (0,0), (-1,-1), 24),
        ("TOPPADDING",    (0,0), (0,0),   36),   # top of first row
        ("BOTTOMPADDING", (0,0), (0,0),   6),
        ("TOPPADDING",    (0,1), (0,1),   4),
        ("BOTTOMPADDING", (0,1), (0,1),   4),
        ("TOPPADDING",    (0,2), (0,2),   4),
        ("BOTTOMPADDING", (0,2), (0,2),   4),
        ("TOPPADDING",    (0,3), (0,3),   8),
        ("BOTTOMPADDING", (0,3), (0,3),   36),  # bottom of last row
    ]
    banner_tbl = Table(banner_rows, colWidths=[17*cm])
    banner_tbl.setStyle(TableStyle(banner_padding))
    story.append(banner_tbl)
    story.append(_spacer(0.8))

    # 3. KPI boxes below the banner
    pct_bg = _pct_color(pct_global)
    S_KPI_VAL = _sty("KV", fontSize=16, fontName="Helvetica-Bold",
                     textColor=RL_BLUE, alignment=TA_CENTER, leading=20)
    S_KPI_LAB = _sty("KL", fontSize=8,  fontName="Helvetica",
                     textColor=colors.HexColor("#555555"), alignment=TA_CENTER, leading=11)
    S_KPI_PCT_VAL = _sty("KPV", fontSize=16, fontName="Helvetica-Bold",
                          textColor=RL_WHITE, alignment=TA_CENTER, leading=20)
    S_KPI_PCT_LAB = _sty("KPL", fontSize=8,  fontName="Helvetica",
                          textColor=colors.HexColor("#DDEEFF"), alignment=TA_CENTER, leading=11)

    kpi_rows = [[
        Table([[Paragraph(f"{total_pal:,}", S_KPI_VAL)],
               [Paragraph("Pallets Activos", S_KPI_LAB)]], colWidths=[3.2*cm]),
        Table([[Paragraph(f"{total_m2:,.0f} m²", S_KPI_VAL)],
               [Paragraph("M² Ocupados", S_KPI_LAB)]], colWidths=[3.2*cm]),
        Table([[Paragraph(f"{cap_region:,} m²", S_KPI_VAL)],
               [Paragraph("Capacidad Total", S_KPI_LAB)]], colWidths=[3.2*cm]),
        Table([[Paragraph(f"{pct_global}%", S_KPI_PCT_VAL)],
               [Paragraph("% Ocupación", S_KPI_PCT_LAB)]], colWidths=[3.2*cm]),
        Table([[Paragraph(f"{disponible:,.0f} m²", S_KPI_VAL)],
               [Paragraph("Disponible", S_KPI_LAB)]], colWidths=[3.2*cm]),
    ]]
    kpi_tbl = Table(kpi_rows, colWidths=[3.2*cm]*5)
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (1,0),  colors.HexColor("#EBF0F7")),
        ("BACKGROUND",    (2,0), (2,0),  colors.HexColor("#EBF0F7")),
        ("BACKGROUND",    (3,0), (3,0),  pct_bg),
        ("BACKGROUND",    (4,0), (4,0),  colors.HexColor("#EBF0F7")),
        ("BOX",           (0,0), (-1,-1), 0.5, RL_ACCENT),
        ("INNERGRID",     (0,0), (-1,-1), 0.3, colors.HexColor("#CCDDEE")),
        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
    ]))
    story.append(kpi_tbl)
    story.append(PageBreak())

    # ══════════════════════════════════════════════════════════════════════════
    # PAGE 2: RESUMEN DE OCUPACIÓN
    # ══════════════════════════════════════════════════════════════════════════
    story.append(Paragraph("1. Resumen de Ocupación por Crossdock", S_H1))
    story.append(_hr())

    # Build occupancy fig
    ocp = []
    for xd in xdocks_plot:
        cap    = CAPACIDADES[xd]
        m2_ocp = round(df_plot[df_plot[xdock_col]==xd]["M2"].sum(), 2)
        pct    = m2_ocp / cap if cap > 0 else 0
        ocp.append({"Ciudad": CIUDAD_MAP.get(xd,xd), "pct": round(pct*100,1),
                    "M2_ocp": m2_ocp, "cap": cap, "disp": round(cap-m2_ocp,1), "xd": xd})
    df_ocp = pd.DataFrame(ocp).sort_values("pct", ascending=True)

    fig_ocp = go.Figure(go.Bar(
        x=df_ocp["pct"], y=df_ocp["Ciudad"], orientation="h",
        marker_color=[_color_pct(p/100) for p in df_ocp["pct"]],
        text=[f"{p}%" for p in df_ocp["pct"]], textposition="outside",
    ))
    for v, lbl, clr in [(70,"70%","#E67E22"),(90,"90%","#C0392B"),(100,"100%","#7B241C")]:
        fig_ocp.add_vline(x=v, line_dash="dot", line_color=clr,
                          annotation_text=lbl, annotation_font_color=clr)
    fig_ocp.update_layout(
        title="", xaxis_title="% Ocupación", yaxis_title="",
        plot_bgcolor="white", paper_bgcolor="white",
        height=350, margin=dict(l=10,r=80,t=10,b=30),
        font=dict(family="Helvetica", color="#1A3A6B"),
        xaxis=dict(range=[0, max(df_ocp["pct"].max()*1.18, 115)]),
    )
    story.append(_fig_to_image(fig_ocp, 16, 7.5))
    story.append(_spacer(0.3))

    # ── Per-crossdock analysis table ──────────────────────────────────────────
    story.append(Paragraph("Análisis por Crossdock", S_H2))
    tbl_hdr = ["Ciudad", "Cap. m²", "M² Ocup.", "% Ocup.", "Disponible", "Status"]
    tbl_rows = [tbl_hdr]
    for row_d in ocp:
        xd    = row_d["xd"]
        pct_v = row_d["pct"]
        st_lbl = ("SATURADO" if pct_v>100 else "CRÍTICO" if pct_v>90
                  else "ALERTA" if pct_v>70 else "NORMAL")
        tbl_rows.append([
            row_d["Ciudad"],
            f"{row_d['cap']:,}",
            f"{row_d['M2_ocp']:,.0f}",
            f"{pct_v:.1f}%",
            f"{row_d['disp']:,.0f}",
            st_lbl,
        ])
    ocp_tbl = Table(tbl_rows, colWidths=[3.2*cm, 2.2*cm, 2.2*cm, 2.2*cm, 2.4*cm, 2.4*cm])

    def _status_color(s):
        if "SATURADO" in s: return RL_DRED
        if "CRÍTICO"  in s: return RL_RED
        if "ALERTA"   in s: return RL_AMBER
        return RL_GREEN

    ts = [
        ("BACKGROUND",   (0,0),(-1,0), RL_BLUE),
        ("TEXTCOLOR",    (0,0),(-1,0), RL_WHITE),
        ("FONTNAME",     (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0),(-1,-1), 8),
        ("ALIGN",        (0,0),(-1,-1), "CENTER"),
        ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
        ("ROWHEIGHT",    (0,0),(-1,-1), 0.6*cm),
        ("BOX",          (0,0),(-1,-1), 0.5, RL_ACCENT),
        ("INNERGRID",    (0,0),(-1,-1), 0.3, colors.HexColor("#CCDDEE")),
        ("TOPPADDING",   (0,0),(-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1), 4),
    ]
    for ri, row_d in enumerate(ocp[: ], 1):
        bg = colors.HexColor("#EBF0F7") if ri%2==0 else colors.white
        ts.append(("BACKGROUND", (0,ri),(-2,ri), bg))
        pct_v = row_d["pct"]
        ts.append(("BACKGROUND", (-1,ri),(-1,ri), _status_color(
            "SATURADO" if pct_v>100 else "CRÍTICO" if pct_v>90 else "ALERTA" if pct_v>70 else "NORMAL")))
        ts.append(("TEXTCOLOR",  (-1,ri),(-1,ri), RL_WHITE))
        ts.append(("FONTNAME",   (-1,ri),(-1,ri), "Helvetica-Bold"))

    ocp_tbl.setStyle(TableStyle(ts))
    story.append(ocp_tbl)
    story.append(_spacer(0.4))

    # Global analysis paragraph
    sat = [r for r in ocp if r["pct"]>100]
    crit= [r for r in ocp if 90<r["pct"]<=100]
    ok  = [r for r in ocp if r["pct"]<=70]
    analysis_lines = []
    if sat:
        analysis_lines.append(
            f"<b>Saturación detectada:</b> {', '.join(r['Ciudad'] for r in sat)} superan el 100% de capacidad. "
            "Esto puede indicar capturas sin confirmación de salida o errores de registro — se recomienda auditoria inmediata."
        )
    if crit:
        analysis_lines.append(
            f"<b>Nivel crítico:</b> {', '.join(r['Ciudad'] for r in crit)} están entre el 90-100%. "
            "Priorizar coordinar salidas antes de nuevas recepciones."
        )
    if not sat and not crit:
        analysis_lines.append("El inventario global se encuentra en niveles manejables. Continuar monitoreo semanal.")
    if ok:
        analysis_lines.append(
            f"{', '.join(r['Ciudad'] for r in ok)} tienen capacidad ampliamente disponible."
        )
    for line in analysis_lines:
        story.append(Paragraph(line, S_BODY))

    story.append(PageBreak())

    # ══════════════════════════════════════════════════════════════════════════
    # PAGE 3: M² y MATERIAL
    # ══════════════════════════════════════════════════════════════════════════
    story.append(Paragraph("2. Análisis de M² y Tipo de Material", S_H1))
    story.append(_hr())

    # ── M² por crossdock: grouped bars M² ocupados vs capacidad ─────────────
    # Using df_ocp already computed from df_plot (clean data only, no consolidados)
    # Sort by % ocupacion descending so most critical crossdocks appear first
    df_ocp_sorted = df_ocp.sort_values("pct", ascending=False)

    fig_m2 = go.Figure()
    fig_m2.add_trace(go.Bar(
        x=df_ocp_sorted["Ciudad"], y=df_ocp_sorted["M2_ocp"],
        name="M² Ocupados", marker_color=GASO_LIGHT,
        text=[f"{v:.0f}" for v in df_ocp_sorted["M2_ocp"]],
        textposition="outside",
    ))
    fig_m2.add_trace(go.Bar(
        x=df_ocp_sorted["Ciudad"], y=df_ocp_sorted["cap"],
        name="Capacidad Total", marker_color="#D5E8F5",
        text=[f"{v:.0f}" for v in df_ocp_sorted["cap"]],
        textposition="outside",
    ))
    fig_m2.update_layout(
        barmode="group", title="", xaxis_title="", yaxis_title="m²",
        plot_bgcolor="white", paper_bgcolor="white",
        height=320, margin=dict(l=10, r=10, t=20, b=70),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        font=dict(family="Helvetica", color="#1A3A6B"),
        xaxis=dict(tickangle=-25),
    )
    story.append(_fig_to_image(fig_m2, 17, 7))

    # Contextual paragraph based on actual numbers
    saturados = [r["Ciudad"] for _, r in df_ocp_sorted.iterrows() if r["pct"] > 100]
    en_alerta  = [r["Ciudad"] for _, r in df_ocp_sorted.iterrows() if 70 < r["pct"] <= 100]
    m2_txt = (
        f"La capacidad total de la región es <b>{cap_region:,} m²</b> y "
        f"actualmente se utilizan <b>{total_m2:,.0f} m²</b> ({pct_global}% de ocupación). "
    )
    if saturados:
        m2_txt += (f"<b>{', '.join(saturados)}</b> superan su capacidad — "
                   "se requiere revisión y despacho prioritario. ")
    if en_alerta:
        m2_txt += f"<b>{', '.join(en_alerta)}</b> operan en zona de alerta."
    story.append(Paragraph(m2_txt, S_BODY))
    story.append(_spacer(0.3))

    # ── Pie: distribución de M² por tipo de material (no pallets) ─────────
    # Using M² so the pie reflects actual space consumption, not unit count
    mat_m2 = df_plot.groupby(mat_col)["M2"].sum().reset_index()
    mat_m2.columns = ["Tipo", "M2"]
    mat_m2 = mat_m2[mat_m2["M2"] > 0].sort_values("M2", ascending=False)

    fig_pie = px.pie(
        mat_m2, names="Tipo", values="M2", hole=0.42,
        color_discrete_sequence=[GASO_BLUE, GASO_ACCENT, "#E67E22",
                                  "#8E44AD", "#1E8449", "#C0392B", "#F39C12"],
    )
    fig_pie.update_traces(
        textinfo="percent+label",
        textfont=dict(size=9),
        hovertemplate="<b>%{label}</b><br>%{value:.0f} m²<br>%{percent}<extra></extra>",
    )
    fig_pie.update_layout(
        plot_bgcolor="white", paper_bgcolor="white",
        showlegend=False, height=300,
        margin=dict(l=20, r=20, t=10, b=10),
        font=dict(family="Helvetica", color="#1A3A6B"),
    )

    # ── Heatmap: M² por crossdock × tipo de material ─────────────────────
    heat = df_plot.groupby([xdock_col, mat_col])["M2"].sum().reset_index()
    hp   = heat.pivot(index=xdock_col, columns=mat_col, values="M2").fillna(0)
    # Sort rows by total M2 descending (most occupied first)
    hp["_total"] = hp.sum(axis=1)
    hp = hp.sort_values("_total", ascending=False).drop(columns=["_total"])
    # Map xdock codes to city names on the Y axis
    hp.index = [CIUDAD_MAP.get(x, x) for x in hp.index]
    n_rows = len(hp)

    fig_heat = px.imshow(
        hp.round(0), text_auto=".0f", aspect="auto",
        color_continuous_scale=[[0, "#EBF5FB"], [0.35, GASO_ACCENT], [1, GASO_BLUE]],
        labels=dict(x="Tipo de Material", y="Ciudad", color="M²"),
    )
    fig_heat.update_layout(
        plot_bgcolor="white", paper_bgcolor="white",
        height=max(300, n_rows * 48),
        margin=dict(l=100, r=20, t=30, b=80),
        font=dict(family="Helvetica", size=10, color="#1A3A6B"),
        coloraxis_showscale=False,
        xaxis=dict(side="bottom", tickangle=-35, tickfont=dict(size=9),
                   title="", automargin=True),
        yaxis=dict(tickfont=dict(size=10), title="", automargin=True),
    )
    fig_heat.update_traces(
        textfont=dict(size=9),
        # White text on dark cells, dark on light
        texttemplate="%{z:.0f}",
    )

    story.append(Paragraph("Distribución de M² por Tipo de Material", S_H2))
    story.append(_fig_to_image(fig_pie, 11, 6))
    story.append(_spacer(0.3))
    story.append(Paragraph("M² por Crossdock y Tipo de Material", S_H2))
    heat_h_cm = max(6, n_rows * 1.0)
    story.append(_fig_to_image(fig_heat, 17, heat_h_cm))

    top_mat = mat_m2.iloc[0]["Tipo"] if len(mat_m2) else "N/A"
    top_m2  = mat_m2.iloc[0]["M2"] if len(mat_m2) else 0
    top_pct = round(top_m2 / total_m2 * 100, 1) if total_m2 > 0 else 0
    story.append(Paragraph(
        f"El tipo de material con mayor consumo de espacio es <b>{top_mat}</b> "
        f"con <b>{top_m2:,.0f} m²</b> ({top_pct}% del total ocupado). "
        "El mapa muestra qué tipos de material concentran más espacio en cada crossdock, "
        "lo que permite priorizar despachos por tipo.",
        S_BODY))

    story.append(PageBreak())

    # ══════════════════════════════════════════════════════════════════════════
    # PAGE 4: CARRIER y ANÁLISIS INDIVIDUAL
    # ══════════════════════════════════════════════════════════════════════════
    story.append(Paragraph("3. Análisis por Carrier y Crossdock", S_H1))
    story.append(_hr())

    car_grp = df_plot.groupby(carrier_col)[pallet_col].sum().reset_index()
    car_grp.columns = ["Carrier", "Pallets"]
    fig_car = px.bar(car_grp.sort_values("Pallets", ascending=False),
                     x="Carrier", y="Pallets", text="Pallets",
                     color="Carrier",
                     color_discrete_sequence=[GASO_BLUE, GASO_ACCENT, "#E67E22", "#8E44AD"])
    fig_car.update_traces(textposition="outside")
    fig_car.update_layout(showlegend=False, plot_bgcolor="white", paper_bgcolor="white",
                          height=300, margin=dict(l=10,r=10,t=10,b=40),
                          font=dict(family="Helvetica", color="#1A3A6B"),
                          xaxis_title="", yaxis_title="Pallets")
    story.append(_fig_to_image(fig_car, 16, 6.5))

    top_carrier = car_grp.sort_values("Pallets", ascending=False).iloc[0]["Carrier"] if len(car_grp) else "N/A"
    top_car_pal = int(car_grp.sort_values("Pallets", ascending=False).iloc[0]["Pallets"]) if len(car_grp) else 0
    story.append(Paragraph(
        f"El carrier con mayor volumen de inventario es <b>{top_carrier}</b> con <b>{top_car_pal:,} pallets</b>. "
        "Un alto volumen de un solo carrier puede representar un riesgo operativo "
        "si existen retrasos en sus salidas programadas.",
        S_BODY))
    story.append(_spacer(0.4))

    # ── Individual crossdock analysis ─────────────────────────────────────────
    story.append(Paragraph("4. Análisis Individual por Crossdock", S_H1))
    story.append(_hr())

    for row_d in sorted(ocp, key=lambda r: r["pct"], reverse=True):
        xd     = row_d["xd"]
        ciudad = row_d["Ciudad"]
        cap    = row_d["cap"]
        m2_ocp = row_d["M2_ocp"]
        pct_v  = row_d["pct"]
        pal    = int(df_plot[df_plot[xdock_col]==xd][pallet_col].sum())

        tipo_dist    = df_plot[df_plot[xdock_col]==xd].groupby(cols["tipo_pallet"])[pallet_col].sum().to_dict()
        carrier_dist = df_plot[df_plot[xdock_col]==xd].groupby(carrier_col)[pallet_col].sum().to_dict()

        analysis_text = _smart_analysis(xd, pct_v/100, pal, m2_ocp, cap, tipo_dist, carrier_dist)
        status_lbl    = ("🔴 SATURADO" if pct_v>100 else "🔴 CRÍTICO" if pct_v>90
                         else "🟡 ALERTA" if pct_v>70 else "🟢 NORMAL")
        st_color = (_pct_color(pct_v/100))

        block = [
            Paragraph(f"{ciudad}  —  {status_lbl}", S_H2),
            Table([[
                Paragraph(f"<b>Pallets:</b> {pal:,}", S_BOLD),
                Paragraph(f"<b>M² Ocup.:</b> {m2_ocp:,.0f}", S_BOLD),
                Paragraph(f"<b>Capacidad:</b> {cap:,} m²", S_BOLD),
                Paragraph(f"<b>Ocupación:</b> {pct_v:.1f}%", S_BOLD),
            ]], colWidths=[3.8*cm]*4,
               style=TableStyle([
                   ("BACKGROUND",(0,0),(-1,-1), colors.HexColor("#F4F6F9")),
                   ("BOX",(0,0),(-1,-1),0.5,RL_ACCENT),
                   ("INNERGRID",(0,0),(-1,-1),0.2,colors.HexColor("#CCDDEE")),
                   ("ALIGN",(0,0),(-1,-1),"CENTER"),
                   ("TOPPADDING",(0,0),(-1,-1),5),
                   ("BOTTOMPADDING",(0,0),(-1,-1),5),
               ])),
            _spacer(0.2),
            Paragraph(analysis_text, S_BODY),
            _spacer(0.2),
            _hr(),
        ]
        story.append(KeepTogether(block))

    # ══════════════════════════════════════════════════════════════════════════
    # LAST PAGE: SITIOS CONSOLIDADOS + FOOTER
    # ══════════════════════════════════════════════════════════════════════════
    story.append(PageBreak())
    story.append(Paragraph("5. Sitios Consolidados", S_H1))
    story.append(_hr())

    df_cons_region = df_consol[df_consol[xdock_col].isin(xdocks_plot)] if len(df_consol) > 0 else df_consol
    n_consol_r = len(df_cons_region)
    story.append(Paragraph(
        f"Se identificaron <b>{n_consol_r} sitios consolidados</b> (pallets con valor 0) "
        "que corresponden a material pequeño ubicado sobre otra tarima. "
        "Estos registros <b>no impactan el cálculo de capacidad ni de m²</b> "
        "pero se mantienen en el inventario para trazabilidad.",
        S_BODY))

    if n_consol_r > 0:
        grp_c = df_cons_region.groupby([xdock_col, carrier_col]).size().reset_index(name="Sitios")
        grp_c["Ciudad"] = grp_c[xdock_col].map(CIUDAD_MAP)
        cons_data  = [["Ciudad", "Carrier", "Sitios Consolidados"]]
        for _, r in grp_c.iterrows():
            cons_data.append([r["Ciudad"], r[carrier_col], str(r["Sitios"])])
        cons_tbl = Table(cons_data, colWidths=[5*cm, 5*cm, 5*cm])
        cons_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0), RL_BLUE),
            ("TEXTCOLOR", (0,0),(-1,0), RL_WHITE),
            ("FONTNAME",  (0,0),(-1,0), "Helvetica-Bold"),
            ("FONTSIZE",  (0,0),(-1,-1), 9),
            ("ALIGN",     (0,0),(-1,-1), "CENTER"),
            ("ROWHEIGHT", (0,0),(-1,-1), 0.65*cm),
            ("BOX",       (0,0),(-1,-1), 0.5, RL_ACCENT),
            ("INNERGRID", (0,0),(-1,-1), 0.3, colors.HexColor("#CCDDEE")),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#EBF0F7")]),
        ]))
        story.append(cons_tbl)

    # Footer note
    story.append(_spacer(1))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#CCDDEE")))
    story.append(Paragraph(
        f"Reporte generado automáticamente por el sistema IN-OUT de Gaso Comunicaciones  |  {fecha_str}  |  Confidencial",
        _sty("Footer", fontSize=7, textColor=colors.HexColor("#AAAAAA"),
             fontName="Helvetica", alignment=TA_CENTER, spaceBefore=6)
    ))

    doc.build(story)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
#  CROSSDOCK DEEP ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────

def build_crossdock_excel(df_raw_full, xdock_name, cols):
    """
    Build a professional analyst-grade Excel workbook for a single crossdock.
    Uses Excel formulas throughout so all metrics update when data changes.
    df_raw_full: the FULL raw dataframe (before pipeline filtering) so we
                 have SALIDA rows too for balance analysis.
    """
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    import pandas as pd
    import numpy as np
    from unidecode import unidecode

    ciudad   = CIUDAD_MAP.get(xdock_name, xdock_name)
    fecha_hoy = datetime.date.today().strftime("%d/%m/%Y")

    # ── Palette ──────────────────────────────────────────────────────────────
    C_BLUE   = "1A3A6B"
    C_LIGHT  = "2E6DB4"
    C_ACCENT = "4A90D9"
    C_WHITE  = "FFFFFF"
    C_LGRAY  = "EBF0F7"
    C_MGRAY  = "D5E3F5"
    C_GREEN  = "1E8449"
    C_AMBER  = "E67E22"
    C_RED    = "C0392B"
    C_DRED   = "7B241C"
    C_GOLD   = "F5C518"
    C_PURPLE = "6C3483"

    def _fill(hex_c):   return PatternFill("solid", fgColor=hex_c)
    def _font(bold=False, color=C_WHITE, size=10, name="Calibri"):
        return Font(name=name, bold=bold, color=color, size=size)
    def _ctr(wrap=True):
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    def _lft():         return Alignment(horizontal="left",  vertical="center")
    def _rgt():         return Alignment(horizontal="right", vertical="center")

    _th = Side(style="thin",   color="BBBBBB")
    _md = Side(style="medium", color=C_BLUE)
    BORDER  = Border(left=_th,  right=_th,  top=_th,  bottom=_th)
    BORDER_M= Border(left=_md,  right=_md,  top=_md,  bottom=_md)

    def _hdr_cell(ws, row, col, val, bg=C_BLUE, fg=C_WHITE, sz=10, bold=True):
        c = ws.cell(row=row, column=col, value=val)
        c.font = _font(bold=bold, color=fg, size=sz)
        c.fill = _fill(bg); c.border = BORDER; c.alignment = _ctr()
        return c

    def _data_cell(ws, row, col, val, alt=False, fmt=None, align="left"):
        c = ws.cell(row=row, column=col, value=val)
        c.font   = _font(bold=False, color="222222")
        c.fill   = _fill(C_LGRAY if alt else C_WHITE)
        c.border = BORDER
        c.alignment = _ctr() if align == "center" else (_rgt() if align == "right" else _lft())
        if fmt: c.number_format = fmt
        return c

    def _title_block(ws, title, subtitle, n_cols=20):
        ws.row_dimensions[1].height = 36
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 16
        end = get_column_letter(n_cols)
        ws.merge_cells(f"A1:{end}1")
        c = ws["A1"]
        c.value = f"GASO COMUNICACIONES  –  {ciudad.upper()}"
        c.font  = _font(bold=True, size=14, color=C_WHITE)
        c.fill  = _fill(C_BLUE); c.alignment = _ctr()
        ws.merge_cells(f"A2:{end}2")
        c = ws["A2"]
        c.value = title
        c.font  = _font(bold=True, size=11, color=C_WHITE)
        c.fill  = _fill(C_LIGHT); c.alignment = _ctr()
        ws.merge_cells(f"A3:{end}3")
        c = ws["A3"]
        c.value = f"{subtitle}   |   {xdock_name}   |   Generado: {fecha_hoy}"
        c.font  = _font(bold=False, size=9, color="555555")
        c.fill  = _fill("F4F6F9"); c.alignment = _ctr()

    # ── Prepare data ─────────────────────────────────────────────────────────
    # Filter for this crossdock (raw data has XDOCK aliases, normalize)
    def _match_xdock(val):
        if not val: return False
        v = str(val).strip()
        return v == xdock_name or XDOCK_ALIASES.get(norm(v), v) == xdock_name

    df_xd = df_raw_full[df_raw_full["XDOCK"].apply(_match_xdock)].copy()

    # Parse dates
    df_xd["_FECHA_ING"] = pd.to_datetime(df_xd["FECHA DE INGRESO"], errors="coerce")
    df_xd["_FECHA_SAL"] = pd.to_datetime(df_xd["FECHA DE SALIDA"],  errors="coerce")
    df_xd["_MES_ING"]   = df_xd["_FECHA_ING"].dt.to_period("M").astype(str)
    df_xd["_AÑO_ING"]   = df_xd["_FECHA_ING"].dt.year
    df_xd["_NO_PAL"]    = pd.to_numeric(df_xd["NO. DE PALLET"], errors="coerce").fillna(0)
    df_xd["_DIAS"]      = pd.to_numeric(df_xd["DIAS INV."], errors="coerce")
    df_xd["_DIAS_DUR"]  = pd.to_numeric(df_xd["DIAS QUE DURO EN INVENTARIO"], errors="coerce")

    # Active (inventario) vs Salidas
    def _is_salida(row):
        es = str(row.get("ESTATUS SALIDA", "")).strip().upper()
        fd = row.get("FECHA DE SALIDA", None)
        # A record is a SALIDA only when ESTATUS SALIDA = "SALIDA"
        # AND has a real (non-null) FECHA DE SALIDA
        fd_valid = (fd is not None and
                    not (isinstance(fd, float) and pd.isna(fd)) and
                    str(fd).strip() not in ("", "None", "NaT", "nan"))
        return es == "SALIDA" and fd_valid

    df_xd["_ES_SALIDA"] = df_xd.apply(_is_salida, axis=1)
    df_activo  = df_xd[~df_xd["_ES_SALIDA"]].copy()
    df_salidas = df_xd[df_xd["_ES_SALIDA"]].copy()

    carriers = sorted(df_xd["CARRIER"].dropna().unique())
    tipos    = sorted(df_xd["TIPO DE MATERIAL"].fillna("SIN CLASIFICAR").unique())
    clasifs  = sorted(df_xd["CLASIFICACION DE MATERIAL"].fillna("OTRO").unique())
    cap      = CAPACIDADES.get(xdock_name, 0)

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1: DATOS BASE (raw filtered data — dynamic source for all formulas)
    # ══════════════════════════════════════════════════════════════════════════
    ws_data = wb.active
    ws_data.title = "DATOS"
    ws_data.sheet_view.showGridLines = False

    # Write full raw data (normalized)
    export_cols = [
        "CARRIER","XDOCK","FECHA DE INGRESO","ESTATUS","CLASIFICACION DE MATERIAL",
        "TIPO DE MATERIAL","ID SITIO","NOMBRE DE SITIO","NO. DE PALLET","TIPO DE PALLET",
        "DESCRIPCION MATERIAL","PROYECTO","ESTATUS SALIDA","DIAS INV.",
        "FECHA DE SALIDA","PALLETS SALIDA","EXISTENCIA REAL","DIAS QUE DURO EN INVENTARIO",
        "NOMBRE ASP","OBSERVACIONES",
    ]
    export_cols = [c for c in export_cols if c in df_xd.columns]

    _title_block(ws_data, f"BASE DE DATOS – {ciudad}", "Fuente dinámica para todos los análisis", n_cols=len(export_cols))
    ws_data.row_dimensions[5].height = 28
    for ci, h in enumerate(export_cols, 1):
        _hdr_cell(ws_data, 5, ci, h, bg=C_BLUE)
        ws_data.column_dimensions[get_column_letter(ci)].width = max(len(h)+2, 14)

    df_export = df_xd[export_cols].copy()
    df_export["FECHA DE INGRESO"] = df_export["FECHA DE INGRESO"].apply(
        lambda x: x.strftime("%Y-%m-%d") if hasattr(x, "strftime") else x)
    df_export["FECHA DE SALIDA"]  = df_export["FECHA DE SALIDA"].apply(
        lambda x: x.strftime("%Y-%m-%d") if hasattr(x, "strftime") and pd.notna(x) else x)         if "FECHA DE SALIDA" in df_export.columns else df_export.get("FECHA DE SALIDA", "")

    for ri, (_, row) in enumerate(df_export.iterrows()):
        er  = 6 + ri
        alt = ri % 2 == 1
        for ci, col in enumerate(export_cols, 1):
            val = row[col]
            if pd.isna(val): val = None
            _data_cell(ws_data, er, ci, val, alt=alt)

    ws_data.freeze_panes = "A6"
    ws_data.auto_filter.ref = f"A5:{get_column_letter(len(export_cols))}5"
    n_data_rows = len(df_export)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2: RESUMEN EJECUTIVO (KPIs + Excel formulas referencing DATOS)
    # ══════════════════════════════════════════════════════════════════════════
    ws_res = wb.create_sheet("RESUMEN EJECUTIVO")
    ws_res.sheet_view.showGridLines = False
    _title_block(ws_res, f"RESUMEN EJECUTIVO – {ciudad}",
                 "Indicadores clave calculados dinámicamente desde la base de datos", n_cols=14)

    # -- Column definitions for DATOS sheet references --
    # Find column indices in export_cols
    def _col(name):
        try: return get_column_letter(export_cols.index(name) + 1)
        except: return "A"

    c_carrier   = _col("CARRIER")
    c_pal       = _col("NO. DE PALLET")
    c_est_sal   = _col("ESTATUS SALIDA")
    c_dias      = _col("DIAS INV.")
    c_tipo_mat  = _col("TIPO DE MATERIAL")
    c_clasif    = _col("CLASIFICACION DE MATERIAL")
    c_exist     = _col("EXISTENCIA REAL")
    c_pal_sal   = _col("PALLETS SALIDA")
    data_range  = f"DATOS!{c_carrier}6:{c_carrier}{5+n_data_rows}"

    # KPI Section
    def _kpi_block(ws, row, col, label, formula, fmt="0", note=""):
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row,   end_column=col+2)
        c = ws.cell(row=row, column=col, value=label)
        c.font = _font(bold=True, size=9, color=C_BLUE); c.alignment = _lft()

        ws.merge_cells(start_row=row+1, start_column=col,
                       end_row=row+1,   end_column=col+2)
        c2 = ws.cell(row=row+1, column=col, value=formula)
        c2.font   = _font(bold=True, size=18, color=C_BLUE)
        c2.fill   = _fill(C_LGRAY); c2.alignment = _ctr(); c2.border = BORDER_M
        c2.number_format = fmt
        if note:
            ws.merge_cells(start_row=row+2, start_column=col,
                           end_row=row+2,   end_column=col+2)
            c3 = ws.cell(row=row+2, column=col, value=note)
            c3.font = _font(bold=False, size=8, color="888888"); c3.alignment = _ctr()
        return c2

    # Row 5: section title
    ws_res.merge_cells("A5:N5")
    t = ws_res["A5"]; t.value = "▌ INDICADORES CLAVE DE DESEMPEÑO"
    t.font = _font(bold=True, size=11, color=C_BLUE); t.alignment = _lft()
    ws_res.row_dimensions[5].height = 22

    # Set column widths
    for ci in range(1, 15):
        ws_res.column_dimensions[get_column_letter(ci)].width = 12

    # KPI row 1 (row 6-8): Pallets activos | M2 | % Ocupación | Capacidad | Días prom
    r = 6
    # Total pallets activos (EXISTENCIA REAL = 1)
    pal_formula = f'=SUMPRODUCT((DATOS!{c_exist}6:DATOS!{c_exist}{5+n_data_rows}=1)*(ISNUMBER(DATOS!{c_pal}6:DATOS!{c_pal}{5+n_data_rows})*IFERROR(VALUE(DATOS!{c_pal}6:DATOS!{c_pal}{5+n_data_rows}),0)))'
    _kpi_block(ws_res, r, 1,  "📦 PALLETS ACTIVOS",        pal_formula, "0,0")
    _kpi_block(ws_res, r, 5,  "📊 TOTAL REGISTROS",        f'=COUNTA(DATOS!{c_carrier}6:DATOS!{c_carrier}{5+n_data_rows})', "0,0")
    _kpi_block(ws_res, r, 9,  "📤 SALIDAS TOTALES",        f'=COUNTIF(DATOS!{c_est_sal}6:DATOS!{c_est_sal}{5+n_data_rows},"SALIDA")', "0,0")
    _kpi_block(ws_res, r, 12, "🔄 DEVOLUCIONES",           f'=COUNTIF(DATOS!ESTATUS6:DATOS!ESTATUS{5+n_data_rows},"DEVOLUCION")', "0,0")

    ws_res.row_dimensions[6].height = 16
    ws_res.row_dimensions[7].height = 36
    ws_res.row_dimensions[8].height = 14

    r = 10
    _kpi_block(ws_res, r, 1,  "⏱️ DÍAS PROM. EN INVENTARIO", f'=IFERROR(AVERAGEIF(DATOS!{c_exist}6:DATOS!{c_exist}{5+n_data_rows},1,DATOS!{c_dias}6:DATOS!{c_dias}{5+n_data_rows}),0)', "0.0")
    _kpi_block(ws_res, r, 5,  "⏳ MÁXIMO DÍAS EN INV.",      f'=IFERROR(MAXIFS(DATOS!{c_dias}6:DATOS!{c_dias}{5+n_data_rows},DATOS!{c_exist}6:DATOS!{c_exist}{5+n_data_rows},1),0)',     "0")
    _kpi_block(ws_res, r, 9,  "📅 DÍAS PROM. ROTACIÓN",     f'=IFERROR(AVERAGE(DATOS!{"DIAS QUE DURO EN INVENTARIO" and _col("DIAS QUE DURO EN INVENTARIO")}6:DATOS!{_col("DIAS QUE DURO EN INVENTARIO")}{5+n_data_rows}),0)', "0.0", "solo items con salida")
    _kpi_block(ws_res, r, 12, "🏭 CAPACIDAD (m²)",          cap, "0,0")

    ws_res.row_dimensions[10].height = 16
    ws_res.row_dimensions[11].height = 36
    ws_res.row_dimensions[12].height = 14

    # Separator
    r = 14
    ws_res.merge_cells(f"A{r}:N{r}")
    t = ws_res[f"A{r}"]; t.value = "▌ BALANCE POR CARRIER"
    t.font = _font(bold=True, size=11, color=C_BLUE); t.alignment = _lft()
    ws_res.row_dimensions[r].height = 22

    # Balance table per carrier with Excel formulas
    r = 15
    hdrs_bal = ["CARRIER","ENTRADAS","SALIDAS","BALANCE NETO",
                "% SALIDA","PALLETS ACTIVOS","DÍAS PROM INV","DEVOL."]
    for ci, h in enumerate(hdrs_bal, 1):
        _hdr_cell(ws_res, r, ci, h, bg=C_LIGHT)

    carrier_col_idx = export_cols.index("CARRIER") + 1
    exist_col_idx   = export_cols.index("EXISTENCIA REAL") + 1 if "EXISTENCIA REAL" in export_cols else 1
    est_sal_idx     = export_cols.index("ESTATUS SALIDA") + 1 if "ESTATUS SALIDA" in export_cols else 1
    dias_idx        = export_cols.index("DIAS INV.") + 1 if "DIAS INV." in export_cols else 1
    estatus_idx     = export_cols.index("ESTATUS") + 1 if "ESTATUS" in export_cols else 1

    cc = get_column_letter(carrier_col_idx)
    ce = get_column_letter(exist_col_idx)
    ces= get_column_letter(est_sal_idx)
    cd = get_column_letter(dias_idx)
    cst= get_column_letter(estatus_idx)
    data_r = f"6:{5+n_data_rows}"

    for ri, car in enumerate(carriers):
        er  = 16 + ri
        alt = ri % 2 == 1
        entradas = f'=COUNTIF(DATOS!{cc}{data_r.split(":")[0]}:DATOS!{cc}{data_r.split(":")[1]},"{car}")'
        salidas  = f'=COUNTIFS(DATOS!{cc}{data_r.split(":")[0]}:DATOS!{cc}{data_r.split(":")[1]},"{car}",DATOS!{ces}{data_r.split(":")[0]}:DATOS!{ces}{data_r.split(":")[1]},"SALIDA")'
        activos  = f'=COUNTIFS(DATOS!{cc}{data_r.split(":")[0]}:DATOS!{cc}{data_r.split(":")[1]},"{car}",DATOS!{ce}{data_r.split(":")[0]}:DATOS!{ce}{data_r.split(":")[1]},1)'
        devoluc  = f'=COUNTIFS(DATOS!{cc}{data_r.split(":")[0]}:DATOS!{cc}{data_r.split(":")[1]},"{car}",DATOS!{cst}{data_r.split(":")[0]}:DATOS!{cst}{data_r.split(":")[1]},"DEVOLUCION")'
        dias_pr  = f'=IFERROR(AVERAGEIFS(DATOS!{cd}{data_r.split(":")[0]}:DATOS!{cd}{data_r.split(":")[1]},DATOS!{cc}{data_r.split(":")[0]}:DATOS!{cc}{data_r.split(":")[1]},"{car}",DATOS!{ce}{data_r.split(":")[0]}:DATOS!{ce}{data_r.split(":")[1]},1),"-")'

        ent_ref  = get_column_letter(2)
        sal_ref  = get_column_letter(3)
        bal_f    = f"={get_column_letter(2)}{er}-{get_column_letter(3)}{er}"
        pct_f    = f'=IFERROR({get_column_letter(3)}{er}/{get_column_letter(2)}{er},0)'

        vals = [car, entradas, salidas, bal_f, pct_f, activos, dias_pr, devoluc]
        fmts = [None, "0,0", "0,0", "0,0", "0.0%", "0,0", "0.0", "0,0"]
        for ci, (v, f) in enumerate(zip(vals, fmts), 1):
            c = _data_cell(ws_res, er, ci, v, alt=alt)
            if f: c.number_format = f
            if ci == 4:  # Balance neto — color
                c.font = _font(bold=True, color="222222")
            if ci == 5:
                c.number_format = "0.0%"

    # Total row
    tr = 16 + len(carriers)
    _hdr_cell(ws_res, tr, 1, "TOTAL GENERAL", bg=C_BLUE)
    for ci in range(2, 9):
        if ci <= 4:
            c = ws_res.cell(row=tr, column=ci,
                            value=f'=SUM({get_column_letter(ci)}16:{get_column_letter(ci)}{tr-1})')
        elif ci == 5:
            c = ws_res.cell(row=tr, column=ci,
                            value=f'=IFERROR(C{tr}/B{tr},0)')
            c.number_format = "0.0%"
        else:
            c = ws_res.cell(row=tr, column=ci,
                            value=f'=SUM({get_column_letter(ci)}16:{get_column_letter(ci)}{tr-1})')
        c.font = _font(bold=True, color=C_WHITE); c.fill = _fill(C_BLUE)
        c.border = BORDER; c.alignment = _ctr()

    ws_res.row_dimensions[15].height = 28
    for rr in range(16, tr+1):
        ws_res.row_dimensions[rr].height = 22

    # ── Balance por tipo de material (with formulas) ─────────────────────────
    c_tipo_idx = export_cols.index("TIPO DE MATERIAL") + 1 if "TIPO DE MATERIAL" in export_cols else 1
    ct = get_column_letter(c_tipo_idx)

    r = tr + 3
    ws_res.merge_cells(f"A{r}:N{r}")
    t = ws_res[f"A{r}"]; t.value = "▌ DISTRIBUCIÓN POR TIPO DE MATERIAL"
    t.font = _font(bold=True, size=11, color=C_BLUE); t.alignment = _lft()

    r += 1
    mat_hdrs = ["TIPO DE MATERIAL","ENTRADAS TOTALES","EN INVENTARIO","SALIDAS","% ROTACIÓN","DÍAS PROM. INV."]
    for ci, h in enumerate(mat_hdrs, 1):
        _hdr_cell(ws_res, r, ci, h, bg=C_PURPLE)

    for ri, tipo in enumerate(tipos):
        er  = r + 1 + ri
        alt = ri % 2 == 1
        tipo_safe = tipo if tipo else "SIN CLASIFICAR"
        entradas_t = f'=COUNTIF(DATOS!{ct}{data_r.split(":")[0]}:DATOS!{ct}{data_r.split(":")[1]},"{tipo_safe}")'
        activos_t  = f'=COUNTIFS(DATOS!{ct}{data_r.split(":")[0]}:DATOS!{ct}{data_r.split(":")[1]},"{tipo_safe}",DATOS!{ce}{data_r.split(":")[0]}:DATOS!{ce}{data_r.split(":")[1]},1)'
        salidas_t  = f'=COUNTIFS(DATOS!{ct}{data_r.split(":")[0]}:DATOS!{ct}{data_r.split(":")[1]},"{tipo_safe}",DATOS!{ces}{data_r.split(":")[0]}:DATOS!{ces}{data_r.split(":")[1]},"SALIDA")'
        pct_t      = f'=IFERROR(D{er}/B{er},0)'
        dias_t     = f'=IFERROR(AVERAGEIFS(DATOS!{cd}{data_r.split(":")[0]}:DATOS!{cd}{data_r.split(":")[1]},DATOS!{ct}{data_r.split(":")[0]}:DATOS!{ct}{data_r.split(":")[1]},"{tipo_safe}",DATOS!{ce}{data_r.split(":")[0]}:DATOS!{ce}{data_r.split(":")[1]},1),"-")'
        vals = [tipo_safe, entradas_t, activos_t, salidas_t, pct_t, dias_t]
        fmts = [None, "0,0", "0,0", "0,0", "0.0%", "0.0"]
        for ci, (v, f) in enumerate(zip(vals, fmts), 1):
            c = _data_cell(ws_res, er, ci, v, alt=alt)
            if f: c.number_format = f
        ws_res.row_dimensions[er].height = 20

    # ── AGING ANALYSIS ───────────────────────────────────────────────────────
    r2 = r + 1 + len(tipos) + 3
    ws_res.merge_cells(f"A{r2}:N{r2}")
    t = ws_res[f"A{r2}"]; t.value = "▌ ANÁLISIS DE ANTIGÜEDAD (AGING) – INVENTARIO ACTIVO"
    t.font = _font(bold=True, size=11, color=C_BLUE); t.alignment = _lft()

    aging_buckets = [
        ("0–15 días",   0,  15),
        ("16–30 días",  16, 30),
        ("31–60 días",  31, 60),
        ("61–90 días",  61, 90),
        ("91–180 días", 91, 180),
        ("+180 días",   181, 99999),
    ]
    r2 += 1
    aging_hdrs = ["RANGO","PALLETS","% DEL TOTAL","RIESGO"]
    bg_aging   = [C_GREEN, C_GREEN, C_AMBER, C_AMBER, C_RED, C_DRED]
    for ci, h in enumerate(aging_hdrs, 1):
        _hdr_cell(ws_res, r2, ci, h, bg=C_BLUE)

    aging_val_rows = []
    for ri, (label, lo, hi) in enumerate(aging_buckets):
        er  = r2 + 1 + ri
        alt = ri % 2 == 1
        hi_cap = min(hi, 99998)
        cnt = f'=COUNTIFS(DATOS!{cd}{data_r.split(":")[0]}:DATOS!{cd}{data_r.split(":")[1]},">={lo}",DATOS!{cd}{data_r.split(":")[0]}:DATOS!{cd}{data_r.split(":")[1]},"<={hi_cap}",DATOS!{ce}{data_r.split(":")[0]}:DATOS!{ce}{data_r.split(":")[1]},1)'
        pct = f'=IFERROR(B{er}/SUM(B{r2+1}:B{r2+6}),0)'
        risk = ["FRESCO","FRESCO","ATENCIÓN","ATENCIÓN","CRÍTICO","URGENTE"][ri]
        for ci, v in enumerate([label, cnt, pct, risk], 1):
            c = _data_cell(ws_res, er, ci, v, alt=alt)
            if ci == 3: c.number_format = "0.0%"
            if ci == 4:
                c.font = _font(bold=True, color=C_WHITE)
                c.fill = _fill(bg_aging[ri])
        ws_res.row_dimensions[er].height = 20
        aging_val_rows.append(er)

    # ── Chart: Aging bar embedded in RESUMEN ─────────────────────────────────
    if aging_val_rows:
        aging_chart = BarChart()
        aging_chart.type = "col"
        aging_chart.title = f"Aging Inventario – {ciudad}"
        aging_chart.style = 10
        aging_chart.y_axis.title = "Pallets"
        aging_chart.x_axis.title = "Rango de días"
        aging_chart.width  = 14
        aging_chart.height = 10

        labels = Reference(ws_res, min_col=1, min_row=r2+1, max_row=r2+6)
        data_c = Reference(ws_res, min_col=2, min_row=r2, max_row=r2+6)
        aging_chart.add_data(data_c, titles_from_data=True)
        aging_chart.set_categories(labels)
        colors_aging = ["1E8449","1E8449","E67E22","E67E22","C0392B","7B241C"]
        for i, ser in enumerate(aging_chart.series):
            ser.graphicalProperties.solidFill = colors_aging[min(i, len(colors_aging)-1)]
        ws_res.add_chart(aging_chart, f"F{r2}")

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3: TENDENCIA MENSUAL
    # ══════════════════════════════════════════════════════════════════════════
    ws_trend = wb.create_sheet("TENDENCIA MENSUAL")
    ws_trend.sheet_view.showGridLines = False
    _title_block(ws_trend, f"TENDENCIA MENSUAL – {ciudad}",
                 "Evolución de entradas, salidas y balance neto mes a mes", n_cols=12)

    # Build monthly aggregation in Python (needed for chart data)
    df_xd["_MES_STR"] = df_xd["_FECHA_ING"].dt.strftime("%Y-%m")
    df_xd["_MES_STR"] = df_xd["_MES_STR"].fillna("DESCONOCIDO")
    meses_valid = df_xd[df_xd["_MES_STR"] != "DESCONOCIDO"]["_MES_STR"]
    meses_range = sorted(meses_valid.unique())

    # Monthly table
    r = 5
    trend_hdrs = ["MES"] + [f"ENT.\n{car}" for car in carriers] +                  ["TOTAL ENT."] + [f"SAL.\n{car}" for car in carriers] +                  ["TOTAL SAL.", "BALANCE"]
    for ci, h in enumerate(trend_hdrs, 1):
        _hdr_cell(ws_trend, r, ci, h, bg=C_BLUE)
        ws_trend.column_dimensions[get_column_letter(ci)].width = 12
    ws_trend.column_dimensions["A"].width = 11

    for ri, mes in enumerate(meses_range):
        er  = 6 + ri
        alt = ri % 2 == 1
        ws_trend.cell(row=er, column=1, value=mes).alignment = _ctr()
        ws_trend.cell(row=er, column=1).fill = _fill(C_LGRAY if alt else C_WHITE)

        ent_vals = []
        for ci, car in enumerate(carriers, 2):
            n = len(df_xd[(df_xd["_MES_STR"]==mes) & (df_xd["CARRIER"]==car)])
            c = ws_trend.cell(row=er, column=ci, value=n)
            c.fill = _fill(C_LGRAY if alt else C_WHITE); c.alignment = _ctr()
            c.number_format = "0"; c.border = BORDER
            ent_vals.append(n)

        tot_ent_col = 2 + len(carriers)
        tot_ent = f'=SUM(B{er}:{get_column_letter(tot_ent_col-1)}{er})'
        c = ws_trend.cell(row=er, column=tot_ent_col, value=tot_ent)
        c.font = _font(bold=True, color="222222")
        c.fill = _fill(C_MGRAY); c.alignment = _ctr(); c.border = BORDER
        c.number_format = "0"

        sal_vals = []
        for ci2, car in enumerate(carriers, tot_ent_col+1):
            n = len(df_xd[(df_xd["_MES_STR"]==mes) &
                          (df_xd["CARRIER"]==car) &
                          (df_xd["_ES_SALIDA"]==True)])
            c = ws_trend.cell(row=er, column=ci2, value=n)
            c.fill = _fill(C_LGRAY if alt else C_WHITE); c.alignment = _ctr()
            c.number_format = "0"; c.border = BORDER
            sal_vals.append(n)

        tot_sal_col = tot_ent_col + 1 + len(carriers)
        tot_sal = f'=SUM({get_column_letter(tot_ent_col+1)}{er}:{get_column_letter(tot_sal_col-1)}{er})'
        c = ws_trend.cell(row=er, column=tot_sal_col, value=tot_sal)
        c.font = _font(bold=True, color="222222")
        c.fill = _fill(C_MGRAY); c.alignment = _ctr(); c.border = BORDER
        c.number_format = "0"

        bal_col = tot_sal_col + 1
        bal_f = f'={get_column_letter(tot_ent_col)}{er}-{get_column_letter(tot_sal_col)}{er}'
        c = ws_trend.cell(row=er, column=bal_col, value=bal_f)
        c.font = _font(bold=True, color="222222")
        c.fill = _fill("D5F5E3" if True else "FADBD8")
        c.number_format = "0"; c.border = BORDER; c.alignment = _ctr()

        ws_trend.row_dimensions[er].height = 20

    n_mes_rows = len(meses_range)
    last_mes_row = 5 + n_mes_rows

    # Totals
    tr_mes = last_mes_row + 1
    _hdr_cell(ws_trend, tr_mes, 1, "TOTALES", bg=C_BLUE)
    for ci in range(2, len(trend_hdrs)+1):
        c = ws_trend.cell(row=tr_mes, column=ci,
                          value=f'=SUM({get_column_letter(ci)}6:{get_column_letter(ci)}{last_mes_row})')
        c.font = _font(bold=True, color=C_WHITE); c.fill = _fill(C_BLUE)
        c.border = BORDER; c.alignment = _ctr(); c.number_format = "0"

    ws_trend.row_dimensions[5].height = 28
    ws_trend.freeze_panes = "B6"

    # Chart: Line chart entradas vs salidas mensual
    if n_mes_rows >= 2:
        line_chart = LineChart()
        line_chart.title = f"Tendencia Mensual – {ciudad}"
        line_chart.style = 10
        line_chart.y_axis.title = "Unidades"
        line_chart.x_axis.title = "Mes"
        line_chart.width = 22; line_chart.height = 12

        cats = Reference(ws_trend, min_col=1, min_row=6, max_row=5+n_mes_rows)
        ent_data = Reference(ws_trend, min_col=tot_ent_col, min_row=5, max_row=5+n_mes_rows)
        sal_data = Reference(ws_trend, min_col=tot_sal_col, min_row=5, max_row=5+n_mes_rows)

        line_chart.add_data(ent_data, titles_from_data=True)
        line_chart.add_data(sal_data, titles_from_data=True)
        line_chart.set_categories(cats)
        line_chart.series[0].graphicalProperties.line.solidFill = C_BLUE
        line_chart.series[0].graphicalProperties.line.width     = 20000
        if len(line_chart.series) > 1:
            line_chart.series[1].graphicalProperties.line.solidFill = C_RED
            line_chart.series[1].graphicalProperties.line.width     = 20000

        ws_trend.add_chart(line_chart, f"A{tr_mes+3}")

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4: ANÁLISIS POR CARRIER (one section per carrier)
    # ══════════════════════════════════════════════════════════════════════════
    ws_car = wb.create_sheet("ANÁLISIS CARRIER")
    ws_car.sheet_view.showGridLines = False
    _title_block(ws_car, f"ANÁLISIS POR CARRIER – {ciudad}",
                 "Detalle de entradas, salidas y distribución por tipo de material por carrier", n_cols=14)

    r_car = 5
    for car in carriers:
        df_car = df_xd[df_xd["CARRIER"] == car]
        df_car_act = df_car[~df_car["_ES_SALIDA"]]
        df_car_sal = df_car[df_car["_ES_SALIDA"]]

        # Carrier header
        ws_car.row_dimensions[r_car].height = 24
        ws_car.merge_cells(f"A{r_car}:N{r_car}")
        c = ws_car[f"A{r_car}"]
        c.value = f"▌ CARRIER: {car}"
        c.font  = _font(bold=True, size=12, color=C_WHITE)
        c.fill  = _fill(C_BLUE if car == "AT&T" else C_PURPLE)
        c.alignment = _lft()
        r_car += 1

        # Mini KPIs
        kpi_data = [
            ("Entradas totales", len(df_car)),
            ("En inventario",    len(df_car_act)),
            ("Salidas",          len(df_car_sal)),
            ("% Rotación",       f"{len(df_car_sal)/len(df_car)*100:.1f}%" if len(df_car)>0 else "0%"),
            ("Días prom. inv.",  f"{df_car_act['_DIAS'].median():.0f}" if len(df_car_act)>0 else "-"),
        ]
        for ci, (lbl, val) in enumerate(kpi_data, 1):
            ws_car.cell(row=r_car,   column=ci*2-1, value=lbl).font  = _font(bold=True, size=8, color=C_BLUE)
            c = ws_car.cell(row=r_car+1, column=ci*2-1, value=val)
            c.font = _font(bold=True, size=14, color=C_BLUE)
            c.fill = _fill(C_LGRAY); c.border = BORDER_M; c.alignment = _ctr()
            ws_car.merge_cells(start_row=r_car+1, start_column=ci*2-1,
                               end_row=r_car+1,   end_column=ci*2)
        r_car += 3

        # By tipo de material
        _hdr_cell(ws_car, r_car, 1, "TIPO DE MATERIAL", bg=C_LIGHT)
        _hdr_cell(ws_car, r_car, 2, "ENTRADAS",         bg=C_LIGHT)
        _hdr_cell(ws_car, r_car, 3, "EN INVENTARIO",    bg=C_LIGHT)
        _hdr_cell(ws_car, r_car, 4, "SALIDAS",          bg=C_LIGHT)
        _hdr_cell(ws_car, r_car, 5, "% ROTACIÓN",       bg=C_LIGHT)
        _hdr_cell(ws_car, r_car, 6, "DÍAS PROM. INV.",  bg=C_LIGHT)
        r_car += 1

        tipos_car = sorted(df_car["TIPO DE MATERIAL"].fillna("SIN CLASIFICAR").unique())
        for ri, tipo in enumerate(tipos_car):
            alt   = ri % 2 == 1
            df_t  = df_car[df_car["TIPO DE MATERIAL"].fillna("SIN CLASIFICAR") == tipo]
            df_ta = df_t[~df_t["_ES_SALIDA"]]
            df_ts = df_t[df_t["_ES_SALIDA"]]
            n_ent = len(df_t); n_act = len(df_ta); n_sal = len(df_ts)
            pct_r = n_sal/n_ent if n_ent > 0 else 0
            dias_ = df_ta["_DIAS"].median() if len(df_ta) > 0 else None

            vals = [tipo, n_ent, n_act, n_sal, pct_r, dias_]
            fmts = [None, "0", "0", "0", "0.0%", "0.0"]
            for ci, (v, f) in enumerate(zip(vals, fmts), 1):
                c = _data_cell(ws_car, r_car, ci, v, alt=alt)
                if f: c.number_format = f
            ws_car.row_dimensions[r_car].height = 20
            r_car += 1

        # By clasificacion
        r_car += 1
        _hdr_cell(ws_car, r_car, 1, "CLASIFICACIÓN",      bg=C_ACCENT)
        _hdr_cell(ws_car, r_car, 2, "ENTRADAS",            bg=C_ACCENT)
        _hdr_cell(ws_car, r_car, 3, "EN INVENTARIO",       bg=C_ACCENT)
        _hdr_cell(ws_car, r_car, 4, "TARIMA MÁS FRECUENTE",bg=C_ACCENT)
        r_car += 1
        clasifs_car = sorted(df_car["CLASIFICACION DE MATERIAL"].fillna("OTRO").unique())
        for ri, cl in enumerate(clasifs_car):
            alt  = ri % 2 == 1
            df_c = df_car[df_car["CLASIFICACION DE MATERIAL"].fillna("OTRO") == cl]
            df_ca= df_c[~df_c["_ES_SALIDA"]]
            if len(df_c) > 0 and "TIPO DE PALLET" in df_c.columns:
                top_pallet = df_c["TIPO DE PALLET"].value_counts().index[0] if len(df_c["TIPO DE PALLET"].dropna()) > 0 else "-"
            else:
                top_pallet = "-"
            vals = [cl, len(df_c), len(df_ca), top_pallet]
            for ci, v in enumerate(vals, 1):
                _data_cell(ws_car, r_car, ci, v, alt=alt)
            ws_car.row_dimensions[r_car].height = 20
            r_car += 1

        # Pie chart for this carrier
        if len(tipos_car) >= 2:
            pie = PieChart()
            pie.title  = f"{car} – Entradas por Tipo de Material"
            pie.style  = 10
            pie.width  = 12; pie.height = 8

            tipo_counts = df_car["TIPO DE MATERIAL"].fillna("SIN CLASIFICAR").value_counts()
            # Write a small helper range for the chart
            chart_row_start = r_car + 1
            ws_car.cell(row=chart_row_start, column=8, value="Tipo").font = _font(bold=True, color=C_BLUE)
            ws_car.cell(row=chart_row_start, column=9, value="N").font   = _font(bold=True, color=C_BLUE)
            for pi, (tp, cnt) in enumerate(tipo_counts.items(), 1):
                ws_car.cell(row=chart_row_start+pi, column=8, value=tp)
                ws_car.cell(row=chart_row_start+pi, column=9, value=cnt)

            pie_labels = Reference(ws_car, min_col=8,
                                   min_row=chart_row_start+1,
                                   max_row=chart_row_start+len(tipo_counts))
            pie_data   = Reference(ws_car, min_col=9,
                                   min_row=chart_row_start,
                                   max_row=chart_row_start+len(tipo_counts))
            pie.add_data(pie_data, titles_from_data=True)
            pie.set_categories(pie_labels)
            ws_car.add_chart(pie, f"J{r_car}")

        r_car += 3

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 5: TOP SITES & ASP
    # ══════════════════════════════════════════════════════════════════════════
    ws_top = wb.create_sheet("TOP SITIOS & ASP")
    ws_top.sheet_view.showGridLines = False
    _title_block(ws_top, f"TOP SITIOS Y ASP – {ciudad}",
                 "Sitios y transportistas con más movimientos", n_cols=10)

    # Top 20 sites by entries
    top_sites = df_xd.groupby("ID SITIO").agg(
        Entradas=("CARRIER","count"),
        En_Inventario=("EXISTENCIA REAL", lambda x: (x==1).sum()),
        Dias_Prom=("_DIAS", "median"),
    ).sort_values("Entradas", ascending=False).head(20).reset_index()

    r = 5
    ws_top.merge_cells(f"A{r}:J{r}")
    t = ws_top[f"A{r}"]; t.value = "▌ TOP 20 SITIOS POR ENTRADAS"
    t.font = _font(bold=True, size=11, color=C_BLUE); t.alignment = _lft()
    r += 1
    site_hdrs = ["ID SITIO","NOMBRE SITIO","ENTRADAS","EN INV.","DÍAS PROM.","CARRIER PRINCIPAL"]
    for ci, h in enumerate(site_hdrs, 1):
        _hdr_cell(ws_top, r, ci, h, bg=C_LIGHT)
        ws_top.column_dimensions[get_column_letter(ci)].width = 18
    r += 1
    for ri, row_d in enumerate(top_sites.itertuples(), 0):
        alt = ri % 2 == 1
        site_id = row_d._1
        site_name = df_xd[df_xd["ID SITIO"]==site_id]["NOMBRE DE SITIO"].dropna().iloc[0]                     if len(df_xd[df_xd["ID SITIO"]==site_id]["NOMBRE DE SITIO"].dropna())>0 else ""
        top_car = df_xd[df_xd["ID SITIO"]==site_id]["CARRIER"].value_counts().index[0]                   if len(df_xd[df_xd["ID SITIO"]==site_id])>0 else ""
        vals = [site_id, site_name, row_d.Entradas,
                int(row_d.En_Inventario), round(row_d.Dias_Prom,0) if pd.notna(row_d.Dias_Prom) else None, top_car]
        for ci, v in enumerate(vals, 1):
            _data_cell(ws_top, r, ci, v, alt=alt)
        ws_top.row_dimensions[r].height = 20
        r += 1

    # Top ASP
    r += 2
    ws_top.merge_cells(f"A{r}:J{r}")
    t = ws_top[f"A{r}"]; t.value = "▌ TOP ASP (TRANSPORTISTAS) – SALIDAS"
    t.font = _font(bold=True, size=11, color=C_BLUE); t.alignment = _lft()
    r += 1

    if "NOMBRE ASP" in df_xd.columns:
        top_asp = df_xd[df_xd["_ES_SALIDA"]].groupby("NOMBRE ASP").agg(
            Salidas=("CARRIER","count"),
        ).sort_values("Salidas", ascending=False).head(15).reset_index()

        asp_hdrs = ["NOMBRE ASP","SALIDAS REALIZADAS","% DEL TOTAL"]
        for ci, h in enumerate(asp_hdrs, 1):
            _hdr_cell(ws_top, r, ci, h, bg=C_GREEN)
        r += 1
        tot_sal_asp = len(df_xd[df_xd["_ES_SALIDA"]])
        for ri, row_d in enumerate(top_asp.itertuples(), 0):
            alt = ri % 2 == 1
            pct_asp = row_d.Salidas / tot_sal_asp if tot_sal_asp > 0 else 0
            vals = [row_d._1, row_d.Salidas, pct_asp]
            fmts = [None, "0", "0.0%"]
            for ci, (v, f) in enumerate(zip(vals, fmts), 1):
                c = _data_cell(ws_top, r, ci, v, alt=alt)
                if f: c.number_format = f
            ws_top.row_dimensions[r].height = 20
            r += 1

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 6: ESTADÍSTICAS Y TENDENCIA (regression + stats)
    # ══════════════════════════════════════════════════════════════════════════
    ws_stat = wb.create_sheet("ESTADÍSTICAS")
    ws_stat.sheet_view.showGridLines = False
    _title_block(ws_stat, f"ESTADÍSTICAS AVANZADAS – {ciudad}",
                 "Análisis de tendencia, regresión lineal y proyecciones", n_cols=12)

    r = 5
    # Regression on monthly data
    df_mes = df_xd[df_xd["_MES_STR"] != "DESCONOCIDO"].groupby("_MES_STR").size().reset_index(name="entradas")
    df_mes = df_mes.sort_values("_MES_STR").reset_index(drop=True)
    df_mes["x"] = range(len(df_mes))

    # Write regression data
    ws_stat.merge_cells(f"A{r}:L{r}")
    t = ws_stat[f"A{r}"]; t.value = "▌ REGRESIÓN LINEAL – ENTRADAS MENSUALES"
    t.font = _font(bold=True, size=11, color=C_BLUE); t.alignment = _lft()
    r += 1

    reg_hdrs = ["MES","N° PERÍODO","ENTRADAS","TENDENCIA (LINREG)","DESV. VS TENDENCIA"]
    for ci, h in enumerate(reg_hdrs, 1):
        _hdr_cell(ws_stat, r, ci, h, bg=C_BLUE)
        ws_stat.column_dimensions[get_column_letter(ci)].width = 18
    r += 1

    reg_data_start = r
    for ri, row_d in df_mes.iterrows():
        er  = r + ri
        alt = ri % 2 == 1
        ws_stat.cell(row=er, column=1, value=row_d["_MES_STR"])
        ws_stat.cell(row=er, column=2, value=int(row_d["x"]))
        ws_stat.cell(row=er, column=3, value=int(row_d["entradas"]))
        # TREND formula using FORECAST or LINEST
        if ri >= 2:
            trend_f = (f'=IFERROR(FORECAST(B{er},$C${reg_data_start}:$C${reg_data_start+len(df_mes)-1},'
                       f'$B${reg_data_start}:$B${reg_data_start+len(df_mes)-1}),"-")')
        else:
            trend_f = f'=C{er}'
        c_trend = ws_stat.cell(row=er, column=4, value=trend_f)
        c_trend.number_format = "0.0"
        dev_f = f'=IFERROR(C{er}-D{er},"-")'
        c_dev = ws_stat.cell(row=er, column=5, value=dev_f)
        c_dev.number_format = "0"

        for ci in range(1, 6):
            c = ws_stat.cell(row=er, column=ci)
            c.fill   = _fill(C_LGRAY if alt else C_WHITE)
            c.border = BORDER
            if ci > 1: c.alignment = _ctr()
        ws_stat.row_dimensions[er].height = 20

    reg_data_end = r + len(df_mes) - 1

    # Stats block
    stat_r = reg_data_end + 3
    ws_stat.merge_cells(f"A{stat_r}:L{stat_r}")
    t = ws_stat[f"A{stat_r}"]; t.value = "▌ ESTADÍSTICAS DESCRIPTIVAS – DÍAS EN INVENTARIO (ACTIVOS)"
    t.font = _font(bold=True, size=11, color=C_BLUE); t.alignment = _lft()
    stat_r += 1

    dias_activos = df_activo["_DIAS"].dropna()
    stat_vals = [
        ("Media (días)",      round(dias_activos.mean(),1)       if len(dias_activos) > 0 else 0),
        ("Mediana (días)",    round(dias_activos.median(),1)     if len(dias_activos) > 0 else 0),
        ("Desv. estándar",    round(dias_activos.std(),1)        if len(dias_activos) > 0 else 0),
        ("Percentil 25%",     round(dias_activos.quantile(.25),1)if len(dias_activos) > 0 else 0),
        ("Percentil 75%",     round(dias_activos.quantile(.75),1)if len(dias_activos) > 0 else 0),
        ("Percentil 90%",     round(dias_activos.quantile(.90),1)if len(dias_activos) > 0 else 0),
        ("Máximo (días)",     round(dias_activos.max(),0)        if len(dias_activos) > 0 else 0),
        ("Items >90 días",    int((dias_activos > 90).sum())     if len(dias_activos) > 0 else 0),
        ("Items >180 días",   int((dias_activos > 180).sum())    if len(dias_activos) > 0 else 0),
    ]
    for ci, (lbl, val) in enumerate(stat_vals, 1):
        bg = C_RED if ("90" in lbl or "180" in lbl) and val > 0 else C_LGRAY
        c = ws_stat.cell(row=stat_r,   column=ci, value=lbl)
        c.font = _font(bold=True, size=9, color=C_BLUE); c.alignment = _ctr()
        c2 = ws_stat.cell(row=stat_r+1, column=ci, value=val)
        c2.font = _font(bold=True, size=13, color=C_WHITE if bg==C_RED else C_BLUE)
        c2.fill = _fill(bg); c2.border = BORDER_M; c2.alignment = _ctr()
        c2.number_format = "0.0" if isinstance(val, float) else "0"
        ws_stat.column_dimensions[get_column_letter(ci)].width = 14
    ws_stat.row_dimensions[stat_r].height   = 16
    ws_stat.row_dimensions[stat_r+1].height = 32

    # Proyección próximos 3 meses (simple linear)
    proj_r = stat_r + 4
    ws_stat.merge_cells(f"A{proj_r}:L{proj_r}")
    t = ws_stat[f"A{proj_r}"]; t.value = "▌ PROYECCIÓN PRÓXIMOS 3 MESES (REGRESIÓN LINEAL)"
    t.font = _font(bold=True, size=11, color=C_BLUE); t.alignment = _lft()
    proj_r += 1

    n_mes = len(df_mes)
    for pi in range(3):
        mes_label = (pd.Timestamp.today() + pd.DateOffset(months=pi+1)).strftime("%Y-%m")
        x_proj = n_mes + pi
        proj_f = (f'=IFERROR(FORECAST({x_proj},$C${reg_data_start}:$C${reg_data_end},'
                  f'$B${reg_data_start}:$B${reg_data_end}),"N/D")')
        ws_stat.cell(row=proj_r, column=1, value="MES PROYECTADO").font = _font(bold=True, color=C_BLUE, size=9)
        ws_stat.cell(row=proj_r, column=1).alignment = _ctr()
        ws_stat.cell(row=proj_r+1, column=1, value=mes_label).alignment = _ctr()
        ws_stat.cell(row=proj_r, column=2+pi*2, value="ENTRADAS PROYECTADAS").font = _font(bold=True, color=C_BLUE, size=9)
        ws_stat.cell(row=proj_r, column=2+pi*2).alignment = _ctr()
        c_proj = ws_stat.cell(row=proj_r+1, column=2+pi*2, value=proj_f)
        c_proj.font = _font(bold=True, size=13, color=C_WHITE)
        c_proj.fill = _fill(C_ACCENT); c_proj.border = BORDER_M
        c_proj.alignment = _ctr(); c_proj.number_format = "0"

    # Line chart: trend vs actual
    if len(df_mes) >= 3:
        trend_chart = LineChart()
        trend_chart.title  = f"Tendencia de Entradas – {ciudad}"
        trend_chart.style  = 10
        trend_chart.y_axis.title = "Entradas"
        trend_chart.x_axis.title = "Mes"
        trend_chart.width  = 22; trend_chart.height = 12

        cats_t = Reference(ws_stat, min_col=1, min_row=reg_data_start, max_row=reg_data_end)
        real_d = Reference(ws_stat, min_col=3, min_row=reg_data_start-1, max_row=reg_data_end)
        tren_d = Reference(ws_stat, min_col=4, min_row=reg_data_start-1, max_row=reg_data_end)
        trend_chart.add_data(real_d, titles_from_data=True)
        trend_chart.add_data(tren_d, titles_from_data=True)
        trend_chart.set_categories(cats_t)
        trend_chart.series[0].graphicalProperties.line.solidFill = C_ACCENT
        trend_chart.series[0].graphicalProperties.line.width     = 20000
        if len(trend_chart.series) > 1:
            trend_chart.series[1].graphicalProperties.line.solidFill = C_RED
            trend_chart.series[1].graphicalProperties.line.width     = 15000
            trend_chart.series[1].graphicalProperties.line.dashDot   = "dash"

        ws_stat.add_chart(trend_chart, f"A{proj_r+4}")

    # Save
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_crossdock_pdf(df_raw_full, xdock_name, cols):
    """
    Build an executive PDF report for a single crossdock with smart analysis.
    """
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                     Table, TableStyle, PageBreak, HRFlowable,
                                     KeepTogether)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
    from reportlab.lib.pagesizes import A4
    import pandas as pd, numpy as np

    ciudad    = CIUDAD_MAP.get(xdock_name, xdock_name)
    cap       = CAPACIDADES.get(xdock_name, 0)
    fecha_str = datetime.date.today().strftime("%d de %B de %Y")
    region    = REGION_MAP.get(xdock_name, "")

    def _match_xdock(val):
        if not val: return False
        v = str(val).strip()
        return v == xdock_name or XDOCK_ALIASES.get(norm(v), v) == xdock_name

    df_xd = df_raw_full[df_raw_full["XDOCK"].apply(_match_xdock)].copy()
    df_xd["_FECHA_ING"] = pd.to_datetime(df_xd["FECHA DE INGRESO"], errors="coerce")
    df_xd["_NO_PAL"]    = pd.to_numeric(df_xd["NO. DE PALLET"], errors="coerce").fillna(0)
    df_xd["_DIAS"]      = pd.to_numeric(df_xd["DIAS INV."], errors="coerce")
    df_xd["_MES_STR"]   = df_xd["_FECHA_ING"].dt.strftime("%Y-%m").fillna("DESCONOCIDO")

    def _is_salida(row):
        es = str(row.get("ESTATUS SALIDA","")).strip().upper()
        fd = row.get("FECHA DE SALIDA", None)
        fd_valid = (fd is not None and
                    not (isinstance(fd, float) and pd.isna(fd)) and
                    str(fd).strip() not in ("", "None", "NaT", "nan"))
        return es == "SALIDA" and fd_valid

    df_xd["_ES_SALIDA"] = df_xd.apply(_is_salida, axis=1)
    df_activo  = df_xd[~df_xd["_ES_SALIDA"]]
    df_salidas = df_xd[ df_xd["_ES_SALIDA"]]
    carriers   = sorted(df_xd["CARRIER"].dropna().unique())

    RL_BLUE   = colors.HexColor("#1A3A6B")
    RL_LIGHT  = colors.HexColor("#2E6DB4")
    RL_ACCENT = colors.HexColor("#4A90D9")
    RL_GREEN  = colors.HexColor("#1E8449")
    RL_AMBER  = colors.HexColor("#E67E22")
    RL_RED    = colors.HexColor("#C0392B")
    RL_DRED   = colors.HexColor("#7B241C")
    RL_WHITE  = colors.white
    RL_LGRAY  = colors.HexColor("#EBF0F7")

    def _sty(name, **kw): return ParagraphStyle(name, **kw)
    S_H1   = _sty("H1",  fontSize=14, textColor=RL_BLUE,  fontName="Helvetica-Bold",
                  spaceBefore=12, spaceAfter=4)
    S_H2   = _sty("H2",  fontSize=11, textColor=RL_LIGHT, fontName="Helvetica-Bold",
                  spaceBefore=8,  spaceAfter=3)
    S_BODY = _sty("Body",fontSize=9,  textColor=colors.HexColor("#333333"),
                  fontName="Helvetica", leading=14, alignment=TA_JUSTIFY, spaceAfter=6)
    S_CTR  = _sty("Ctr", fontSize=9,  textColor=colors.HexColor("#333333"),
                  fontName="Helvetica", alignment=TA_CENTER)
    S_FOOT = _sty("Foot",fontSize=7,  textColor=colors.HexColor("#AAAAAA"),
                  fontName="Helvetica", alignment=TA_CENTER, spaceBefore=4)

    def _hr(): return HRFlowable(width="100%", thickness=1, color=RL_ACCENT,
                                  spaceAfter=6, spaceBefore=2)
    def _sp(h=0.3): return Spacer(1, h*cm)

    def _tbl(data, col_widths, style_extra=None):
        base = [
            ("BACKGROUND", (0,0),(-1,0), RL_BLUE),
            ("TEXTCOLOR",  (0,0),(-1,0), RL_WHITE),
            ("FONTNAME",   (0,0),(-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0),(-1,-1), 8),
            ("ALIGN",      (0,0),(-1,-1), "CENTER"),
            ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
            ("ROWHEIGHT",  (0,0),(-1,-1), 0.55*cm),
            ("BOX",        (0,0),(-1,-1), 0.5, RL_ACCENT),
            ("INNERGRID",  (0,0),(-1,-1), 0.25, colors.HexColor("#CCDDEE")),
            ("TOPPADDING", (0,0),(-1,-1), 3),
            ("BOTTOMPADDING",(0,0),(-1,-1), 3),
        ]
        for ri in range(1, len(data)):
            bg = colors.HexColor("#EBF0F7") if ri%2==0 else colors.white
            base.append(("BACKGROUND",(0,ri),(-1,ri), bg))
        if style_extra: base.extend(style_extra)
        t = Table(data, colWidths=col_widths)
        t.setStyle(TableStyle(base))
        return t

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2.5*cm, bottomMargin=2*cm)
    story = []

    # ── COVER ────────────────────────────────────────────────────────────────
    S_BT = _sty("BT", fontSize=24, textColor=RL_WHITE, fontName="Helvetica-Bold",
                alignment=TA_CENTER, leading=30)
    S_BS = _sty("BS", fontSize=12, textColor=colors.HexColor("#C8D8EC"),
                fontName="Helvetica", alignment=TA_CENTER, leading=18)
    S_BD = _sty("BD", fontSize=10, textColor=colors.HexColor("#AABBCC"),
                fontName="Helvetica", alignment=TA_CENTER, leading=14)

    if os.path.exists(LOGO_PATH):
        try:
            from reportlab.platypus import Image as RLImage
            logo = RLImage(LOGO_PATH, width=3.6*cm, height=1.8*cm)
            logo.hAlign = "CENTER"
            story.append(_sp(1))
            story.append(logo)
            story.append(_sp(0.5))
        except Exception: story.append(_sp(3))
    else:
        story.append(_sp(4))

    banner = Table([
        [Paragraph("GASO COMUNICACIONES", S_BT)],
        [Paragraph(f"Reporte Ejecutivo de Crossdock", S_BS)],
        [Paragraph(f"{ciudad.upper()}  ·  {region}", S_BS)],
        [Paragraph(fecha_str, S_BD)],
    ], colWidths=[17*cm])
    banner.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), RL_BLUE),
        ("LEFTPADDING",(0,0),(-1,-1), 24),
        ("RIGHTPADDING",(0,0),(-1,-1), 24),
        ("TOPPADDING",(0,0),(0,0), 36),
        ("BOTTOMPADDING",(0,-1),(-1,-1), 36),
        ("TOPPADDING",(0,1),(-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-2), 4),
    ]))
    story.append(banner)
    story.append(_sp(0.8))

    # KPI cover boxes
    n_ent  = len(df_xd);  n_act = len(df_activo); n_sal = len(df_salidas)
    pct_rot= n_sal/n_ent if n_ent>0 else 0
    dias_p = df_activo["_DIAS"].median() if len(df_activo)>0 else 0

    def _pct_c(p):
        if p>100: return RL_DRED
        if p>90:  return RL_RED
        if p>70:  return RL_AMBER
        return RL_GREEN

    def _kpi_cell(val, lbl, bg=RL_LGRAY, fg=RL_BLUE):
        S_V = _sty("KV", fontSize=15, fontName="Helvetica-Bold",
                   textColor=fg, alignment=TA_CENTER, leading=19)
        S_L = _sty("KL", fontSize=8, fontName="Helvetica",
                   textColor=colors.HexColor("#555555") if bg==RL_LGRAY else RL_WHITE,
                   alignment=TA_CENTER, leading=11)
        inner = Table([[Paragraph(str(val), S_V)],[Paragraph(lbl, S_L)]],
                      colWidths=[3.2*cm])
        inner.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1), bg),
            ("BOX",(0,0),(-1,-1), 0.5, RL_ACCENT),
            ("TOPPADDING",(0,0),(-1,-1), 8),
            ("BOTTOMPADDING",(0,0),(-1,-1), 8),
        ]))
        return inner

    kpi_row = [[
        _kpi_cell(f"{n_ent:,}", "Registros Totales"),
        _kpi_cell(f"{n_act:,}", "En Inventario"),
        _kpi_cell(f"{n_sal:,}", "Salidas Procesadas"),
        _kpi_cell(f"{pct_rot:.0%}", "% Rotación",
                  bg=_pct_c(pct_rot*100 if cap==0 else n_act/cap*100), fg=RL_WHITE),
        _kpi_cell(f"{dias_p:.0f} d", "Días Prom. Inventario"),
    ]]
    kpi_tbl = Table(kpi_row, colWidths=[3.2*cm]*5)
    kpi_tbl.setStyle(TableStyle([
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story.append(kpi_tbl)
    story.append(PageBreak())

    # ── PAGE 2: BALANCE POR CARRIER ──────────────────────────────────────────
    story.append(Paragraph(f"1. Balance de Entradas y Salidas por Carrier", S_H1))
    story.append(_hr())

    bal_hdr = ["Carrier","Entradas","En Inv.","Salidas","% Rotación","Días Prom.","Devoluciones"]
    bal_data = [bal_hdr]
    for car in carriers:
        df_c  = df_xd[df_xd["CARRIER"]==car]
        df_ca = df_c[~df_c["_ES_SALIDA"]]
        df_cs = df_c[df_c["_ES_SALIDA"]]
        df_cd = df_c[df_c.get("ESTATUS","").apply(lambda x: str(x).upper()=="DEVOLUCION") if "ESTATUS" in df_c.columns else pd.Series([False]*len(df_c))]
        pct_c = len(df_cs)/len(df_c) if len(df_c)>0 else 0
        dias_c= df_ca["_DIAS"].median() if len(df_ca)>0 else 0
        bal_data.append([car, f"{len(df_c):,}", f"{len(df_ca):,}", f"{len(df_cs):,}",
                         f"{pct_c:.1%}", f"{dias_c:.0f} días", f"{len(df_cd):,}"])

    story.append(_tbl(bal_data, [3.0*cm,2.0*cm,2.0*cm,2.0*cm,2.2*cm,2.2*cm,2.4*cm]))
    story.append(_sp(0.4))

    # Narrative
    for car in carriers:
        df_c  = df_xd[df_xd["CARRIER"]==car]
        df_ca = df_c[~df_c["_ES_SALIDA"]]
        df_cs = df_c[df_c["_ES_SALIDA"]]
        pct_c = len(df_cs)/len(df_c) if len(df_c)>0 else 0
        dias_c= df_ca["_DIAS"].median() if len(df_ca)>0 else 0
        if pct_c < 0.3:
            txt = (f"<b>{car}</b> muestra una rotación baja del {pct_c:.0%} en {ciudad}. "
                   f"Con {len(df_ca):,} pallets activos y una mediana de {dias_c:.0f} días en inventario, "
                   f"se recomienda revisar el programa de salidas.")
        elif pct_c > 0.7:
            txt = (f"<b>{car}</b> tiene una rotación saludable del {pct_c:.0%}. "
                   f"El crossdock gestiona activamente sus {len(df_ca):,} pallets activos "
                   f"con una permanencia mediana de {dias_c:.0f} días.")
        else:
            txt = (f"<b>{car}</b> registra una rotación del {pct_c:.0%} con {len(df_ca):,} "
                   f"pallets activos y mediana de {dias_c:.0f} días en almacén.")
        story.append(Paragraph(txt, S_BODY))

    story.append(PageBreak())

    # ── PAGE 3: DISTRIBUCIÓN Y AGING ─────────────────────────────────────────
    story.append(Paragraph("2. Distribución por Tipo de Material y Aging", S_H1))
    story.append(_hr())

    mat_hdr = ["Tipo de Material","Entradas","En Inv.","Salidas","% Rotación"]
    mat_data = [mat_hdr]
    for tipo in sorted(df_xd["TIPO DE MATERIAL"].fillna("SIN CLASIFICAR").unique()):
        df_t  = df_xd[df_xd["TIPO DE MATERIAL"].fillna("SIN CLASIFICAR")==tipo]
        df_ta = df_t[~df_t["_ES_SALIDA"]]
        df_ts = df_t[df_t["_ES_SALIDA"]]
        pct_t = len(df_ts)/len(df_t) if len(df_t)>0 else 0
        mat_data.append([tipo, f"{len(df_t):,}", f"{len(df_ta):,}",
                         f"{len(df_ts):,}", f"{pct_t:.1%}"])

    story.append(Paragraph("Distribución por Tipo de Material", S_H2))
    story.append(_tbl(mat_data, [4.5*cm,2.5*cm,2.5*cm,2.5*cm,2.5*cm]))
    story.append(_sp(0.5))

    # Aging
    story.append(Paragraph("Análisis de Antigüedad (Aging) – Inventario Activo", S_H2))
    dias_act = df_activo["_DIAS"].dropna()
    aging_buckets = [("0–15 días",0,15,"FRESCO"),("16–30 días",16,30,"FRESCO"),
                     ("31–60 días",31,60,"ATENCIÓN"),("61–90 días",61,90,"ATENCIÓN"),
                     ("91–180 días",91,180,"CRÍTICO"),("+180 días",181,99999,"URGENTE")]
    aging_hdr  = ["Rango","Pallets","% del Total","Estado"]
    aging_data = [aging_hdr]
    aging_style_extra = []
    for ri, (lbl, lo, hi, risk) in enumerate(aging_buckets, 1):
        cnt = int(((dias_act >= lo) & (dias_act <= min(hi,99998))).sum())
        pct = cnt/len(dias_act) if len(dias_act)>0 else 0
        aging_data.append([lbl, f"{cnt:,}", f"{pct:.1%}", risk])
        risk_color = {"FRESCO":RL_GREEN,"ATENCIÓN":RL_AMBER,
                      "CRÍTICO":RL_RED,"URGENTE":RL_DRED}[risk]
        aging_style_extra.append(("BACKGROUND",(3,ri),(3,ri), risk_color))
        aging_style_extra.append(("TEXTCOLOR", (3,ri),(3,ri), RL_WHITE))
        aging_style_extra.append(("FONTNAME",  (3,ri),(3,ri), "Helvetica-Bold"))

    story.append(_tbl(aging_data, [4*cm,3*cm,3*cm,3*cm], aging_style_extra))

    # Aging narrative
    items_critical = int((dias_act > 90).sum())
    items_urgent   = int((dias_act > 180).sum())
    dias_med       = dias_act.median() if len(dias_act)>0 else 0
    if items_urgent > 0:
        aging_txt = (f"Se detectan <b>{items_urgent} pallets con más de 180 días</b> en inventario — "
                     f"material potencialmente obsoleto que requiere revisión prioritaria. "
                     f"En total, {items_critical} pallets superan los 90 días. "
                     f"La mediana general es de {dias_med:.0f} días.")
    elif items_critical > 0:
        aging_txt = (f"Hay <b>{items_critical} pallets con más de 90 días</b> en {ciudad}. "
                     f"Se recomienda revisión con el carrier correspondiente para programar salidas. "
                     f"La permanencia mediana es de {dias_med:.0f} días.")
    else:
        aging_txt = (f"El inventario de {ciudad} muestra buena rotación: "
                     f"ningún pallet supera los 90 días en almacén. "
                     f"La permanencia mediana es de {dias_med:.0f} días.")
    story.append(_sp(0.3))
    story.append(Paragraph(aging_txt, S_BODY))
    story.append(PageBreak())

    # ── PAGE 4: TENDENCIA Y REGRESIÓN ────────────────────────────────────────
    story.append(Paragraph("3. Tendencia Mensual y Proyección", S_H1))
    story.append(_hr())

    df_mes = df_xd[df_xd["_MES_STR"]!="DESCONOCIDO"].groupby("_MES_STR").agg(
        Entradas=("CARRIER","count"),
        Salidas=("_ES_SALIDA","sum"),
    ).reset_index().sort_values("_MES_STR").tail(18)  # last 18 months

    trend_hdr  = ["Mes","Entradas","Salidas","Balance"]
    trend_data = [trend_hdr]
    for _, row_t in df_mes.iterrows():
        bal = int(row_t["Entradas"]) - int(row_t["Salidas"])
        trend_data.append([row_t["_MES_STR"],
                           f"{int(row_t['Entradas']):,}",
                           f"{int(row_t['Salidas']):,}",
                           f"+{bal:,}" if bal >= 0 else f"{bal:,}"])

    story.append(Paragraph("Últimos 18 meses de movimientos", S_H2))
    story.append(_tbl(trend_data, [4*cm,3*cm,3*cm,3*cm]))
    story.append(_sp(0.4))

    # Linear regression narrative
    if len(df_mes) >= 4:
        import numpy as np
        x  = np.arange(len(df_mes))
        y  = df_mes["Entradas"].values.astype(float)
        m, b = np.polyfit(x, y, 1)
        r2 = 1 - np.sum((y - (m*x+b))**2) / np.sum((y-y.mean())**2) if y.std()>0 else 0
        trend_dir = "creciente" if m > 0.5 else "decreciente" if m < -0.5 else "estable"
        proj_1 = max(0, round(m*(len(df_mes))+b, 0))
        proj_2 = max(0, round(m*(len(df_mes)+1)+b, 0))

        reg_txt = (f"La regresión lineal sobre los últimos {len(df_mes)} meses muestra una "
                   f"tendencia <b>{trend_dir}</b> de <b>{m:+.1f} entradas/mes</b> "
                   f"(R² = {r2:.2f}). "
                   f"{'Un R² superior a 0.6 indica que la tendencia es estadísticamente significativa. ' if r2>0.6 else 'El R² bajo indica variabilidad alta mes a mes, lo que dificulta proyecciones precisas. '}"
                   f"Proyección para los próximos 2 meses: "
                   f"<b>~{proj_1:.0f}</b> y <b>~{proj_2:.0f}</b> entradas respectivamente.")
        story.append(Paragraph(reg_txt, S_BODY))

    story.append(PageBreak())

    # ── PAGE 5: TOP SITIOS ───────────────────────────────────────────────────
    story.append(Paragraph("4. Top Sitios y Transportistas", S_H1))
    story.append(_hr())

    top_sites = df_xd.groupby("ID SITIO").agg(
        Entradas=("CARRIER","count"),
        Carrier=("CARRIER", lambda x: x.value_counts().index[0]),
        Dias=("_DIAS","median"),
    ).sort_values("Entradas", ascending=False).head(15).reset_index()

    site_hdr  = ["ID Sitio","Entradas","Carrier Principal","Días Prom."]
    site_data = [site_hdr]
    for _, r_s in top_sites.iterrows():
        site_data.append([r_s["ID SITIO"], f"{int(r_s['Entradas']):,}",
                          r_s["Carrier"], f"{r_s['Dias']:.0f}" if pd.notna(r_s["Dias"]) else "-"])
    story.append(Paragraph("Top 15 Sitios por Entradas", S_H2))
    story.append(_tbl(site_data, [5*cm,3*cm,4*cm,3*cm]))
    story.append(_sp(0.5))

    if "NOMBRE ASP" in df_xd.columns:
        top_asp = df_xd[df_xd["_ES_SALIDA"]].groupby("NOMBRE ASP").size()                    .sort_values(ascending=False).head(10).reset_index()
        top_asp.columns = ["ASP","Salidas"]
        asp_hdr  = ["Transportista (ASP)","Salidas Realizadas"]
        asp_data = [asp_hdr] + [[r_a["ASP"], f"{int(r_a['Salidas']):,}"] for _, r_a in top_asp.iterrows()]
        story.append(Paragraph("Top 10 Transportistas (ASP) por Salidas", S_H2))
        story.append(_tbl(asp_data, [9*cm, 6*cm]))

    # Footer
    story.append(_sp(1))
    story.append(HRFlowable(width="100%", thickness=0.5,
                             color=colors.HexColor("#CCDDEE")))
    story.append(Paragraph(
        f"Reporte Crossdock {ciudad}  ·  Gaso Comunicaciones  ·  {fecha_str}  ·  Confidencial",
        S_FOOT))

    doc.build(story)
    buf.seek(0)
    return buf

def _color_pct(pct):
    if pct > 1.0:  return "#7B241C"
    if pct > 0.90: return "#C0392B"
    if pct > 0.70: return "#E67E22"
    return "#1E8449"


def make_charts(df_clean, cols):
    xdock_col  = cols["xdock"]
    carrier_col = cols["carrier"]
    mat_col    = cols["tipo_material"]
    pallet_col = cols["no_pallet"]

    # Occupancy bar
    ocp = []
    for xd in CAPACIDADES:
        cap    = CAPACIDADES[xd]
        m2_ocp = df_clean[df_clean[xdock_col] == xd]["M2"].sum()
        pct    = m2_ocp / cap if cap > 0 else 0
        ocp.append({"Ciudad": CIUDAD_MAP.get(xd, xd), "pct": round(pct*100,1),
                    "M2_ocp": round(m2_ocp,1), "cap": cap, "disp": round(cap-m2_ocp,1)})
    df_ocp = pd.DataFrame(ocp).sort_values("pct", ascending=True)

    fig1 = go.Figure(go.Bar(
        x=df_ocp["pct"], y=df_ocp["Ciudad"], orientation="h",
        marker_color=[_color_pct(p/100) for p in df_ocp["pct"]],
        text=[f"{p}%" for p in df_ocp["pct"]], textposition="outside",
        hovertemplate="<b>%{y}</b><br>Ocupación: %{x}%<extra></extra>",
    ))
    for v, label, color in [(70, "70%", "#E67E22"), (90, "90%", "#C0392B"), (100, "100%", "#7B241C")]:
        fig1.add_vline(x=v, line_dash="dot", line_color=color,
                       annotation_text=label, annotation_font_color=color)
    fig1.update_layout(
        title="% Ocupación por Crossdock",
        xaxis=dict(range=[0, max(df_ocp["pct"].max()*1.15, 115)], title="% Ocupación"),
        yaxis_title="", plot_bgcolor="white", paper_bgcolor="white",
        height=380, margin=dict(l=10, r=60, t=40, b=20),
        font=dict(family="Calibri", color=GASO_BLUE),
    )

    # Stacked M2
    fig2 = go.Figure([
        go.Bar(x=df_ocp["Ciudad"], y=df_ocp["M2_ocp"], name="M² Ocupados",
               marker_color=GASO_LIGHT,
               hovertemplate="<b>%{x}</b><br>Ocupados: %{y} m²<extra></extra>"),
        go.Bar(x=df_ocp["Ciudad"], y=df_ocp["disp"], name="Disponible",
               marker_color="#D5E8F5",
               hovertemplate="<b>%{x}</b><br>Disponible: %{y} m²<extra></extra>"),
    ])
    fig2.update_layout(
        barmode="stack", title="M² Ocupados vs Disponibles",
        xaxis_title="", yaxis_title="m²", plot_bgcolor="white", paper_bgcolor="white",
        height=360, margin=dict(l=20, r=20, t=40, b=60),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        font=dict(family="Calibri", color=GASO_BLUE),
    )

    # Material bar
    mat = df_clean.groupby(mat_col)[pallet_col].sum().reset_index()
    mat.columns = ["Tipo", "Pallets"]
    mat = mat.sort_values("Pallets", ascending=False)
    fig3 = px.bar(mat, x="Tipo", y="Pallets", text="Pallets",
                  color="Tipo", color_discrete_sequence=px.colors.qualitative.Bold,
                  title="Pallets por Tipo de Material")
    fig3.update_traces(textposition="outside")
    fig3.update_layout(showlegend=False, plot_bgcolor="white", paper_bgcolor="white",
                       height=340, margin=dict(l=20, r=20, t=40, b=60),
                       font=dict(family="Calibri", color=GASO_BLUE))

    # Carrier pie
    car = df_clean.groupby(carrier_col)[pallet_col].sum().reset_index()
    car.columns = ["Carrier", "Pallets"]
    fig4 = px.pie(car, names="Carrier", values="Pallets", hole=0.45,
                  title="Pallets por Carrier",
                  color_discrete_sequence=[GASO_BLUE, GASO_ACCENT, "#E67E22", "#8E44AD"])
    fig4.update_traces(textinfo="percent+label", pull=[0.03]*len(car))
    fig4.update_layout(plot_bgcolor="white", paper_bgcolor="white",
                       height=340, margin=dict(l=20, r=20, t=40, b=20),
                       font=dict(family="Calibri", color=GASO_BLUE))

    # Heatmap
    heat = df_clean.groupby([xdock_col, mat_col])["M2"].sum().reset_index()
    hp   = heat.pivot(index=xdock_col, columns=mat_col, values="M2").fillna(0)
    hp.index = [CIUDAD_MAP.get(x, x) for x in hp.index]
    fig5 = px.imshow(hp.round(1), text_auto=".0f", aspect="auto",
                     color_continuous_scale=[[0,"#EBF5FB"],[0.5,GASO_ACCENT],[1,GASO_BLUE]],
                     title="Heatmap M²: Crossdock × Tipo de Material")
    fig5.update_layout(plot_bgcolor="white", paper_bgcolor="white",
                       height=380, margin=dict(l=10, r=10, t=50, b=10),
                       font=dict(family="Calibri", color=GASO_BLUE))

    # Region bar
    reg = df_clean.groupby("REGION")["M2"].sum().reset_index()
    fig6 = px.bar(reg, x="REGION", y="M2", text="M2", color="REGION",
                  color_discrete_map={"REGIÓN JOSÉ": GASO_BLUE, "REGIÓN JORGE": GASO_ACCENT},
                  title="M² por Región")
    fig6.update_traces(texttemplate="%{text:.0f} m²", textposition="outside")
    fig6.update_layout(showlegend=False, plot_bgcolor="white", paper_bgcolor="white",
                       height=320, margin=dict(l=20, r=20, t=40, b=20),
                       font=dict(family="Calibri", color=GASO_BLUE))

    # Carrier × Ciudad
    cx = df_clean.groupby([carrier_col, xdock_col])[pallet_col].sum().reset_index()
    cx["Ciudad"] = cx[xdock_col].map(CIUDAD_MAP)
    fig7 = px.bar(cx, x="Ciudad", y=pallet_col, color=carrier_col, barmode="group",
                  color_discrete_sequence=[GASO_BLUE, GASO_ACCENT, "#E67E22"],
                  title="Pallets por Ciudad y Carrier")
    fig7.update_layout(plot_bgcolor="white", paper_bgcolor="white",
                       height=320, margin=dict(l=20, r=20, t=40, b=60),
                       font=dict(family="Calibri", color=GASO_BLUE),
                       xaxis_title="", yaxis_title="Pallets",
                       legend=dict(title="Carrier"))

    return fig1, fig2, fig3, fig4, fig5, fig6, fig7


# ─────────────────────────────────────────────────────────────────────────────
#  STREAMLIT APP
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="GASO – IN-OUT Processor",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(f"""
<style>
html, body, [class*="css"] {{ font-family: 'Calibri', 'Segoe UI', sans-serif; }}
.main-header {{
    background: linear-gradient(135deg, {GASO_BLUE} 0%, {GASO_LIGHT} 100%);
    padding: 1.4rem 2rem; border-radius: 12px; margin-bottom: 1.2rem;
    display: flex; align-items: center; gap: 1.2rem;
}}
.main-header h1 {{ color: white; margin: 0; font-size: 1.7rem; font-weight: 700; }}
.main-header p  {{ color: rgba(255,255,255,.82); margin: 0; font-size: 0.88rem; }}
.kpi-card {{
    background: white; border: 1px solid #DCE8F5;
    border-left: 4px solid {GASO_BLUE};
    border-radius: 10px; padding: .9rem 1rem;
    box-shadow: 0 2px 8px rgba(26,58,107,.08); text-align: center;
}}
.kpi-label {{ color:#666; font-size:.72rem; font-weight:700;
              text-transform:uppercase; letter-spacing:.05em; margin-bottom:.3rem; }}
.kpi-value {{ color:{GASO_BLUE}; font-size:1.75rem; font-weight:700; line-height:1; }}
.kpi-unit  {{ color:#999; font-size:.72rem; margin-top:.2rem; }}
.kpi-red   {{ border-left-color:#C0392B; }}
.kpi-amber {{ border-left-color:#E67E22; }}
.kpi-green {{ border-left-color:#1E8449; }}
.sec-title {{
    color:{GASO_BLUE}; font-size:1.05rem; font-weight:700;
    border-bottom: 2px solid {GASO_ACCENT};
    padding-bottom:.25rem; margin: 1.1rem 0 .7rem 0;
}}
.review-card {{
    background:#FFF8F0; border:1px solid #F0C080;
    border-left:4px solid #E67E22; border-radius:10px;
    padding:1rem 1.2rem; margin-bottom:.8rem;
}}
.review-card h4 {{ color:#7D3C00; margin:0 0 .3rem 0; font-size:.95rem; }}
.review-card p  {{ color:#555; font-size:.82rem; margin:0; }}
.saved-badge {{
    background:#D5F5E3; color:#1E8449; border-radius:20px;
    padding:.15rem .6rem; font-size:.75rem; font-weight:700;
}}
.del-badge {{
    background:#FADBD8; color:#C0392B; border-radius:20px;
    padding:.15rem .6rem; font-size:.75rem; font-weight:700;
}}
.log-box {{
    background:#F0F4F8; border-radius:8px; padding:.7rem 1rem;
    font-size:.8rem; font-family:monospace;
    max-height:200px; overflow-y:auto;
}}
.stDownloadButton > button {{
    background:{GASO_BLUE}; color:white; border:none;
    border-radius:8px; padding:.55rem 1.2rem;
    font-weight:600; font-size:.88rem; width:100%;
    transition:background .2s;
}}
.stDownloadButton > button:hover {{ background:{GASO_LIGHT}; }}
</style>
""", unsafe_allow_html=True)

# ── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(f"""
    <div style="text-align:center;padding:.8rem 0">
      <div style="background:{GASO_BLUE};color:white;border-radius:50%;
                  width:56px;height:56px;line-height:56px;font-size:1.6rem;
                  margin:0 auto .4rem auto;font-weight:700;">G</div>
      <div style="color:{GASO_BLUE};font-weight:700;font-size:.95rem;">GASO COMUNICACIONES</div>
      <div style="color:#888;font-size:.72rem;">IN-OUT Report Processor</div>
    </div><hr style="border-color:#DCE8F5">
    """, unsafe_allow_html=True)

    st.markdown("### 📁 Cargar Archivo")
    uploaded = st.file_uploader(
        "Selecciona el archivo IN-OUT (.xlsx)",
        type=["xlsx"],
        help="Hoja 'IN-OUT' con encabezados en fila 5."
    )

    st.markdown("---")
    st.markdown("### ⚙️ Opciones de Vista")
    show_clean = st.checkbox("Mostrar base limpia", value=False)

    # Decision memory manager
    st.markdown("---")
    st.markdown("### 🧠 Decisiones Guardadas")
    saved = load_decisions()
    if saved:
        st.caption(f"{len(saved)} ID_SITIO(s) con decisión guardada")
        for key, dec in saved.items():
            col_a, col_b = st.columns([3, 1])
            with col_a:
                label = dec.get("xdock", "ELIMINAR") if dec["action"] == "assign" else "🗑 ELIMINAR"
                st.caption(f"**{key}** → {label}")
            with col_b:
                if st.button("✕", key=f"del_{key}", help="Borrar decisión"):
                    delete_decision(key)
                    st.rerun()
    else:
        st.caption("Sin decisiones guardadas aún.")

    st.markdown("---")
    st.markdown(f"""
    <div style="font-size:.7rem;color:#AAA;text-align:center">
    v3.0.0 · Gaso Comunicaciones<br>Procesamiento automático de inventario
    </div>""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="main-header">
  <div style="font-size:2.2rem">📦</div>
  <div>
    <h1>IN-OUT Report Processor</h1>
    <p>Limpieza · Enriquecimiento · Reporte Ejecutivo para Cliente &nbsp;|&nbsp; Gaso Comunicaciones</p>
  </div>
</div>
""", unsafe_allow_html=True)

if not uploaded:
    st.markdown("""
    <div style="text-align:center;padding:3rem;background:#F8FAFC;
                border-radius:12px;border:2px dashed #C8D8EC">
      <div style="font-size:3rem">📂</div>
      <h3 style="color:#1A3A6B">Carga tu archivo IN-OUT para comenzar</h3>
      <p style="color:#888">Formato esperado: <b>Excel (.xlsx)</b> · Hoja <b>IN-OUT</b> · Encabezados en fila 5</p>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ── Process button ────────────────────────────────────────────────────────────
_, col_btn, _ = st.columns([2, 1, 2])
with col_btn:
    process_btn = st.button("🚀 Procesar Archivo", type="primary", use_container_width=True)

if "processed" not in st.session_state:
    st.session_state.processed = False

if process_btn:
    uploaded.seek(0)
    with st.spinner("⚙️ Procesando base de datos..."):
        try:
            saved_dec = load_decisions()
            df_clean, df_consol, df_pending, logs, cols = run_pipeline(uploaded, saved_dec)

            st.session_state.df_raw_full   = df_raw_full
            st.session_state.df_clean      = df_clean
            st.session_state.df_consol     = df_consol
            st.session_state.df_pending    = df_pending
            st.session_state.logs       = logs
            st.session_state.cols       = cols
            st.session_state.processed  = True

            # Pre-build Excel
            uploaded.seek(0)
            excel_buf = build_excel_output(df_clean, df_consol, cols, region_filter='Todas')
            st.session_state.excel_buf = excel_buf

        except Exception as e:
            st.error(f"❌ Error durante el procesamiento: {e}")
            import traceback; st.code(traceback.format_exc())

# ── Results ───────────────────────────────────────────────────────────────────
if st.session_state.processed:
    df_raw_full   = st.session_state.get('df_raw_full', None)
    df_clean      = st.session_state.df_clean
    df_consol     = st.session_state.df_consol
    df_pending    = st.session_state.df_pending
    logs       = st.session_state.logs
    cols       = st.session_state.cols
    # Always ensure excel_buf exists — regenerate if the previous run failed
    if 'excel_buf' not in st.session_state or st.session_state.excel_buf is None:
        with st.spinner("Generando reporte Excel..."):
            st.session_state.excel_buf = build_excel_output(df_clean, df_consol, cols, region_filter='Todas')
    excel_buf = st.session_state.excel_buf

    xdock_col  = cols["xdock"]
    pallet_col = cols["no_pallet"]

    # ── Processing log ────────────────────────────────────────────────────────
    with st.expander("📋 Log de Procesamiento", expanded=False):
        st.markdown('<div class="log-box">' + "<br>".join(logs) + "</div>",
                    unsafe_allow_html=True)


    # ── Filtro de Región ──────────────────────────────────────────────────────
    st.markdown('<div class="sec-title">🗺️ Filtro de Región</div>', unsafe_allow_html=True)
    rf1, rf2, rf3 = st.columns([2, 2, 5])
    with rf1:
        region_filter = st.selectbox(
            "Analizar región",
            ["Todas", "REGIÓN JOSÉ", "REGIÓN JORGE"],
            key="region_filter",
            help="Filtra todos los KPIs, gráficas y tabla de ocupación por región."
        )
    with rf2:
        st.markdown("<br>", unsafe_allow_html=True)
        if region_filter != "Todas":
            xd_in_region = [xd for xd in CAPACIDADES if REGION_MAP.get(xd) == region_filter]
            st.caption(f"**{len(xd_in_region)} crossdocks** en {region_filter}: "
                       + ", ".join(CIUDAD_MAP.get(x,x) for x in xd_in_region))

    # Apply region filter to working dataframe for KPIs / charts / tables
    if region_filter == "Todas":
        df_view = df_clean.copy()
        xdocks_view = list(CAPACIDADES.keys())
    else:
        df_view = df_clean[df_clean["REGION"] == region_filter].copy()
        xdocks_view = [xd for xd in CAPACIDADES if REGION_MAP.get(xd) == region_filter]

    # ── KPIs ──────────────────────────────────────────────────────────────────
    st.markdown('<div class="sec-title">📊 Indicadores Clave de Desempeño</div>',
                unsafe_allow_html=True)

    total_pal  = int(df_view[pallet_col].sum())
    total_m2   = round(df_view["M2"].sum(), 2)
    total_cap  = sum(CAPACIDADES[xd] for xd in xdocks_view)
    pct_global = round(total_m2 / total_cap * 100, 1) if total_cap > 0 else 0
    disponible = round(total_cap - total_m2, 2)
    n_pend     = len(df_pending)
    n_consol_kpi = len(df_consol[df_consol[xdock_col].isin(xdocks_view)] if len(df_consol)>0 else df_consol)

    kpi_cls = "kpi-red" if pct_global > 90 else "kpi-amber" if pct_global > 70 else "kpi-green"

    k1, k2, k3, k4, k5, k6, k7 = st.columns(7)
    kpis_ui = [
        (k1, "Total Pallets",         f"{total_pal:,}",        "unidades",  "kpi-card"),
        (k2, "M² Ocupados",           f"{total_m2:,.1f}",      "m²",        "kpi-card"),
        (k3, "Capacidad Total",       f"{total_cap:,}",        "m²",        "kpi-card"),
        (k4, "% Ocupación Global",    f"{pct_global}%",        "",          f"kpi-card {kpi_cls}"),
        (k5, "M² Disponibles",        f"{disponible:,.1f}",    "m²",        "kpi-card"),
        (k6, "Sitios Consolidados",   f"{n_consol_kpi}",       "registros", "kpi-card" ),
        (k7, "Pendientes Revisión",   f"{n_pend}",             "registros",
         "kpi-card kpi-red" if n_pend > 0 else "kpi-card kpi-green"),
    ]
    for col_w, label, val, unit, cls in kpis_ui:
        with col_w:
            st.markdown(f"""
            <div class="{cls}">
              <div class="kpi-label">{label}</div>
              <div class="kpi-value">{val}</div>
              <div class="kpi-unit">{unit}</div>
            </div>""", unsafe_allow_html=True)

    # ── Occupancy table ───────────────────────────────────────────────────────
    st.markdown('<div class="sec-title">🏭 Ocupación por Crossdock</div>',
                unsafe_allow_html=True)
    ocp_rows = []
    for xd in xdocks_view:
        cap    = CAPACIDADES[xd]
        m2_ocp = round(df_view[df_view[xdock_col] == xd]["M2"].sum(), 2)
        pal    = int(df_view[df_view[xdock_col] == xd][pallet_col].sum())
        pct    = round(m2_ocp / cap * 100, 1) if cap > 0 else 0
        ocp_rows.append({
            "Ciudad":         CIUDAD_MAP.get(xd, xd),
            "Región":         REGION_MAP.get(xd, ""),
            "Capacidad m²":   cap,
            "Pallets":        pal,
            "M² Ocupados":    m2_ocp,
            "% Ocupación":    pct,
            "Disponible m²":  round(cap - m2_ocp, 2),
            "Status":         ("🔴 SATURADO" if pct > 100 else "🔴 CRÍTICO" if pct > 90
                               else "🟡 ALERTA" if pct > 70 else "🟢 NORMAL"),
        })
    df_ocp_t = pd.DataFrame(ocp_rows)

    def _color_ocp_col(col):
        styles = []
        for v in col:
            try:
                p = float(v)
            except (TypeError, ValueError):
                styles.append("")
                continue
            if p > 100:
                styles.append("background-color:#7B241C;color:white;font-weight:700")
            elif p > 90:
                styles.append("background-color:#C0392B;color:white;font-weight:700")
            elif p > 70:
                styles.append("background-color:#E67E22;color:white;font-weight:700")
            else:
                styles.append("background-color:#1E8449;color:white;font-weight:700")
        return styles

    st.dataframe(
        df_ocp_t.style
            .apply(_color_ocp_col, subset=["% Ocupación"])
            .format({"% Ocupación": "{:.1f}%", "M² Ocupados": "{:,.2f}",
                     "Disponible m²": "{:,.2f}", "Capacidad m²": "{:,}"}),
        use_container_width=True, height=390,
    )

    # ── Charts ────────────────────────────────────────────────────────────────
    # ── Sitios Consolidados ───────────────────────────────────────────────────
    n_consol = len(df_consol)
    st.markdown('<div class="sec-title">📌 Sitios Consolidados</div>', unsafe_allow_html=True)
    if n_consol > 0:
        xd_col = cols["xdock"]
        car_col = cols["carrier"]
        # Summary cards row
        consol_by_xd = df_consol.groupby(xd_col).size().reset_index(name="n")
        consol_by_car = df_consol.groupby(car_col).size().reset_index(name="n")
        ca, cb, cc = st.columns(3)
        with ca:
            st.markdown(f'''
            <div class="kpi-card" style="border-left-color:#8E44AD">
              <div class="kpi-label">Total Sitios Consolidados</div>
              <div class="kpi-value" style="color:#8E44AD">{n_consol}</div>
              <div class="kpi-unit">registros · sin impacto en capacidad</div>
            </div>''', unsafe_allow_html=True)
        with cb:
            top_xd = CIUDAD_MAP.get(consol_by_xd.sort_values("n", ascending=False).iloc[0][xd_col], "—") if len(consol_by_xd) > 0 else "—"
            top_n  = int(consol_by_xd["n"].max()) if len(consol_by_xd) > 0 else 0
            st.markdown(f'''
            <div class="kpi-card" style="border-left-color:#8E44AD">
              <div class="kpi-label">Crossdock con más consolidados</div>
              <div class="kpi-value" style="color:#8E44AD;font-size:1.3rem">{top_xd}</div>
              <div class="kpi-unit">{top_n} sitios</div>
            </div>''', unsafe_allow_html=True)
        with cc:
            carriers_c = ", ".join(sorted(df_consol[car_col].dropna().unique()))
            st.markdown(f'''
            <div class="kpi-card" style="border-left-color:#8E44AD">
              <div class="kpi-label">Carriers con consolidados</div>
              <div class="kpi-value" style="color:#8E44AD;font-size:1.1rem">{carriers_c}</div>
              <div class="kpi-unit">&nbsp;</div>
            </div>''', unsafe_allow_html=True)

        st.caption("Estos registros tienen **No. de Pallet = 0** — son materiales pequeños consolidados sobre otra tarima. "
                   "No se contabilizan en pallets ni en M² de capacidad.")

        # Summary table by XDOCK × Carrier
        grp_c = (df_consol.groupby([xd_col, car_col])
                 .agg(Sitios=(cols["id_sitio"], "count"),
                      ID_Sitios=(cols["id_sitio"], lambda x: ", ".join(x.dropna().unique()[:5])))
                 .reset_index())
        grp_c["Ciudad"] = grp_c[xd_col].map(CIUDAD_MAP)
        grp_c = grp_c[['Ciudad', xd_col, car_col, 'Sitios', 'ID_Sitios']]
        grp_c.columns = ['Ciudad', 'XDOCK', 'Carrier', 'Sitios Consolidados', 'ID Sitios (muestra)']
        st.dataframe(grp_c, use_container_width=True, height=min(200 + len(grp_c)*35, 380))
    else:
        st.info("No hay sitios consolidados en esta carga.")

    st.markdown('<div class="sec-title">📈 Dashboard Visual</div>', unsafe_allow_html=True)
    fig1, fig2, fig3, fig4, fig5, fig6, fig7 = make_charts(df_view, cols)

    c1, c2 = st.columns(2)
    with c1: st.plotly_chart(fig1, use_container_width=True)
    with c2: st.plotly_chart(fig2, use_container_width=True)

    c3, c4 = st.columns(2)
    with c3: st.plotly_chart(fig3, use_container_width=True)
    with c4: st.plotly_chart(fig4, use_container_width=True)

    st.plotly_chart(fig5, use_container_width=True)

    c5, c6 = st.columns(2)
    with c5: st.plotly_chart(fig6, use_container_width=True)
    with c6: st.plotly_chart(fig7, use_container_width=True)

    # ── Base Limpia (optional) ────────────────────────────────────────────────
    if show_clean:
        st.markdown('<div class="sec-title">🗂️ Base IN-OUT Limpia</div>',
                    unsafe_allow_html=True)
        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            f_car = st.multiselect("Carrier",        df_clean[cols["carrier"]].dropna().unique().tolist())
        with fc2:
            f_xd  = st.multiselect("XDOCK",          df_clean[xdock_col].dropna().unique().tolist())
        with fc3:
            f_mat = st.multiselect("Tipo Material",  df_clean[cols["tipo_material"]].dropna().unique().tolist())
        df_show = df_clean.copy()
        if f_car: df_show = df_show[df_show[cols["carrier"]].isin(f_car)]
        if f_xd:  df_show = df_show[df_show[xdock_col].isin(f_xd)]
        if f_mat: df_show = df_show[df_show[cols["tipo_material"]].isin(f_mat)]
        st.markdown(f"**{len(df_show):,}** registros mostrados")
        st.dataframe(df_show.reset_index(drop=True), use_container_width=True, height=340)



    # ─────────────────────────────────────────────────────────────────────────
    #  REVISIÓN MANUAL  (with persistent memory)
    # ─────────────────────────────────────────────────────────────────────────
    if n_pend > 0:
        st.markdown(f"""
        <div class="sec-title">
          ⚠️ Revisión Manual de Registros
          &nbsp;<span style="background:#FDEBD0;color:#A04000;border-radius:20px;
          padding:.15rem .6rem;font-size:.78rem;font-weight:700">{n_pend} registros · {df_pending[cols['id_sitio']].nunique()} ID_SITIO únicos</span>
        </div>
        """, unsafe_allow_html=True)

        st.info(
            "Los registros siguientes no pudieron asignarse automáticamente. "
            "Toma una decisión por **ID Sitio** — se guardará y se aplicará "
            "automáticamente la próxima vez que subas el archivo. "
            "Puedes modificar o borrar decisiones en cualquier momento desde la barra lateral."
        )

        # Group by ID_SITIO
        pend_groups = (
            df_pending.groupby(cols["id_sitio"])
            .agg(
                REGISTROS=(cols["id_sitio"], "count"),
                CARRIER=(cols["carrier"], lambda x: ", ".join(x.dropna().unique())),
                NOMBRE_SITIO=(cols["nombre_sitio"], lambda x: ", ".join(x.dropna().unique()[:2])),
                FOLIO_EJEMPLO=(cols["folio"], "first"),
            )
            .reset_index()
        )

        current_decisions = load_decisions()
        any_new = False

        for _, grp_row in pend_groups.iterrows():
            id_s  = str(grp_row[cols["id_sitio"]]).strip()
            key   = id_s.upper()
            regs  = grp_row["REGISTROS"]
            carr  = grp_row["CARRIER"]
            nom   = grp_row["NOMBRE_SITIO"]
            folio = grp_row["FOLIO_EJEMPLO"]

            already = current_decisions.get(key, {})
            badge   = ""
            if already.get("action") == "assign":
                badge = f'<span class="saved-badge">✓ Guardado → {already["xdock"]}</span>'
            elif already.get("action") == "delete":
                badge = '<span class="del-badge">✓ Guardado → ELIMINAR</span>'

            st.markdown(f"""
            <div class="review-card">
              <h4>🔍 ID Sitio: <code>{id_s}</code> &nbsp; {badge}</h4>
              <p>Nombre: <b>{nom}</b> &nbsp;|&nbsp; Carrier: <b>{carr}</b>
                 &nbsp;|&nbsp; Registros afectados: <b>{regs}</b>
                 &nbsp;|&nbsp; Folio ejemplo: <code>{folio}</code></p>
            </div>
            """, unsafe_allow_html=True)

            col_dec, col_xd, col_save = st.columns([2, 3, 1])
            with col_dec:
                action = st.radio(
                    "Acción",
                    ["Asignar XDOCK", "Eliminar registros"],
                    key=f"action_{key}",
                    horizontal=True,
                    index=0 if already.get("action", "assign") == "assign" else 1,
                )
            with col_xd:
                default_xd = already.get("xdock", XDOCK_OPTIONS[0]) if action == "Asignar XDOCK" else XDOCK_OPTIONS[0]
                xd_sel = st.selectbox(
                    "XDOCK a asignar",
                    XDOCK_OPTIONS,
                    index=XDOCK_OPTIONS.index(default_xd) if default_xd in XDOCK_OPTIONS else 0,
                    key=f"xdsel_{key}",
                    disabled=(action == "Eliminar registros"),
                )
            with col_save:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("💾 Guardar", key=f"save_{key}", use_container_width=True):
                    if action == "Asignar XDOCK" and xd_sel == XDOCK_OPTIONS[0]:
                        st.warning("Selecciona un XDOCK válido primero.")
                    else:
                        dec_entry = {
                            "action": "assign" if action == "Asignar XDOCK" else "delete",
                            "xdock":  xd_sel if action == "Asignar XDOCK" else "",
                            "guardado": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
                        }
                        current_decisions[key] = dec_entry
                        save_decisions(current_decisions)
                        any_new = True
                        st.success(f"✅ Decisión guardada para **{id_s}**")

        if any_new:
            st.info("💡 Vuelve a procesar el archivo para aplicar las decisiones guardadas.")

    else:
        st.success("✅ Todos los registros tienen XDOCK asignado. No hay pendientes.")

    # ── Análisis por Crossdock ───────────────────────────────────────────────
    st.markdown('<div class="sec-title">🏭 Análisis Profundo por Crossdock</div>',
                unsafe_allow_html=True)

    all_xdocks = sorted(CAPACIDADES.keys())
    xd_display = {CIUDAD_MAP.get(x, x): x for x in all_xdocks}

    xd_col1, xd_col2 = st.columns([3, 2])
    with xd_col1:
        xd_selected_city = st.selectbox(
            "Selecciona el crossdock a analizar",
            list(xd_display.keys()),
            key="xd_deep_select",
        )
    xd_selected = xd_display[xd_selected_city]

    with xd_col2:
        st.markdown("<br>", unsafe_allow_html=True)
        xd_gen_btn = st.button("📊 Generar Análisis de Crossdock",
                               use_container_width=True, key="xd_gen_btn")

    if xd_gen_btn and df_raw_full is not None:
        with st.spinner(f"Construyendo análisis completo de {xd_selected_city}..."):
            try:
                xd_excel = build_crossdock_excel(df_raw_full, xd_selected, cols)
                xd_pdf   = build_crossdock_pdf(df_raw_full, xd_selected, cols)
                st.session_state.xd_excel      = xd_excel
                st.session_state.xd_pdf        = xd_pdf
                st.session_state.xd_name       = xd_selected
                st.session_state.xd_city       = xd_selected_city
                st.success(f"✅ Análisis de {xd_selected_city} generado correctamente")
            except Exception as e:
                st.error(f"Error generando análisis: {e}")
                import traceback; st.code(traceback.format_exc())
    elif xd_gen_btn and df_raw_full is None:
        st.warning("Procesa el archivo primero.")

    if "xd_excel" in st.session_state:
        xd_city_slug = st.session_state.xd_city.replace(" ","_").upper()
        xd_col_a, xd_col_b = st.columns(2)
        with xd_col_a:
            st.download_button(
                label=f"📥 Excel Analítico – {st.session_state.xd_city}",
                data=st.session_state.xd_excel,
                file_name=f"GASO_CROSSDOCK_{xd_city_slug}_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with xd_col_b:
            st.download_button(
                label=f"📄 PDF Ejecutivo – {st.session_state.xd_city}",
                data=st.session_state.xd_pdf,
                file_name=f"GASO_CROSSDOCK_{xd_city_slug}_{datetime.date.today().strftime('%Y%m%d')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

        # Preview metrics in app
        if df_raw_full is not None:
            import pandas as pd
            def _match(val):
                if not val: return False
                v = str(val).strip()
                return v == st.session_state.xd_name or XDOCK_ALIASES.get(norm(v), v) == st.session_state.xd_name

            df_xd_prev = df_raw_full[df_raw_full["XDOCK"].apply(_match)].copy()
            df_xd_prev["_FECHA_ING"] = pd.to_datetime(df_xd_prev["FECHA DE INGRESO"], errors="coerce")
            df_xd_prev["_DIAS"]      = pd.to_numeric(df_xd_prev.get("DIAS INV.", pd.Series()), errors="coerce")
            df_xd_prev["_EXIST"]     = pd.to_numeric(df_xd_prev.get("EXISTENCIA REAL", pd.Series()), errors="coerce")

            def _is_sal(row):
                es = str(row.get("ESTATUS SALIDA","")).strip().upper()
                fd = row.get("FECHA DE SALIDA", None)
                fd_valid = (fd is not None and
                            not (isinstance(fd, float) and pd.isna(fd)) and
                            str(fd).strip() not in ("", "None", "NaT", "nan"))
                return es == "SALIDA" and fd_valid

            df_xd_prev["_ES_SAL"] = df_xd_prev.apply(_is_sal, axis=1)
            df_act_p = df_xd_prev[~df_xd_prev["_ES_SAL"]]

            with st.expander(f"📋 Vista previa – {st.session_state.xd_city}", expanded=True):
                pm1, pm2, pm3, pm4, pm5 = st.columns(5)
                mets = [
                    (pm1, "Total registros",    f"{len(df_xd_prev):,}"),
                    (pm2, "En inventario",      f"{len(df_act_p):,}"),
                    (pm3, "Salidas procesadas", f"{df_xd_prev['_ES_SAL'].sum():,}"),
                    (pm4, "Días prom. inv.",    f"{df_act_p['_DIAS'].median():.0f}" if len(df_act_p)>0 else "-"),
                    (pm5, ">90 días en inv.",   f"{int((df_act_p['_DIAS'].dropna()>90).sum()):,}"),
                ]
                for col_m, lbl, val in mets:
                    with col_m:
                        st.metric(lbl, val)

                # Carrier breakdown table
                if "CARRIER" in df_xd_prev.columns:
                    car_summ = []
                    for car in sorted(df_xd_prev["CARRIER"].dropna().unique()):
                        df_cc = df_xd_prev[df_xd_prev["CARRIER"]==car]
                        df_ca = df_cc[~df_cc["_ES_SAL"]]
                        pct_r = df_cc["_ES_SAL"].sum()/len(df_cc) if len(df_cc)>0 else 0
                        car_summ.append({
                            "Carrier": car,
                            "Entradas": len(df_cc),
                            "En inventario": len(df_ca),
                            "Salidas": int(df_cc["_ES_SAL"].sum()),
                            "% Rotación": f"{pct_r:.1%}",
                            "Días prom.": f"{df_ca['_DIAS'].median():.0f}" if len(df_ca)>0 else "-",
                        })
                    st.dataframe(pd.DataFrame(car_summ), use_container_width=True, hide_index=True)

    # ── Downloads ─────────────────────────────────────────────────────────────
    st.markdown('<div class="sec-title">⬇️ Descargar Reportes</div>',
                unsafe_allow_html=True)
    fecha_str = datetime.date.today().strftime("%Y%m%d")

    # PDF generation
    st.markdown("**📄 Reporte PDF Ejecutivo**")
    pdf_r1, pdf_r2 = st.columns([3, 2])
    with pdf_r1:
        pdf_region = st.selectbox("Región para el PDF", ["Todas", "REGIÓN JOSÉ", "REGIÓN JORGE"],
                                   key="pdf_region_select")
    with pdf_r2:
        st.markdown("<br>", unsafe_allow_html=True)
        gen_pdf_btn = st.button("🖨️ Generar PDF Ejecutivo", use_container_width=True)

    if gen_pdf_btn:
        with st.spinner("Generando reporte PDF..."):
            try:
                pdf_buf = generate_pdf(df_clean, df_consol, cols, region_filter=pdf_region)
                st.session_state.pdf_buf = pdf_buf
                st.session_state.pdf_region = pdf_region
                st.success("✅ PDF generado correctamente")
            except Exception as e:
                st.error(f"Error generando PDF: {e}")
                import traceback; st.code(traceback.format_exc())

    if "pdf_buf" in st.session_state:
        region_label = st.session_state.get("pdf_region", "Todas").replace(" ","_").replace("Ó","O").replace("É","E")
        st.download_button(
            label="📥 Descargar PDF Ejecutivo",
            data=st.session_state.pdf_buf,
            file_name=f"GASO_REPORTE_EJECUTIVO_{region_label}_{fecha_str}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )

    st.markdown("---")
    st.markdown("**📊 Archivos Excel**")
    xl_r1, xl_r2 = st.columns([3, 2])
    with xl_r1:
        xl_region = st.selectbox("Región para el Excel", ["Todas", "REGIÓN JOSÉ", "REGIÓN JORGE"],
                                  key="xl_region_select")
    with xl_r2:
        st.markdown("<br>", unsafe_allow_html=True)
        gen_xl_btn = st.button("📊 Generar Reporte Excel", use_container_width=True)

    if gen_xl_btn:
        with st.spinner("Generando reporte Excel..."):
            try:
                xl_buf_new = build_excel_output(df_clean, df_consol, cols, region_filter=xl_region)
                st.session_state.xl_buf_custom    = xl_buf_new
                st.session_state.xl_region_custom = xl_region
                st.success("✅ Excel generado correctamente")
            except Exception as e:
                st.error(f"Error generando Excel: {e}")
                import traceback; st.code(traceback.format_exc())

    # Show download button once generated (or use pre-built full version)
    xl_download_buf    = st.session_state.get("xl_buf_custom", excel_buf)
    xl_download_region = st.session_state.get("xl_region_custom", "Todas")
    xl_region_slug     = xl_download_region.replace(" ","_").replace("Ó","O").replace("É","E").replace("Á","A")

    d1, d2, d3 = st.columns(3)
    with d1:
        st.download_button(
            label="📥 Reporte Completo Excel",
            data=xl_download_buf,
            file_name=f"GASO_REPORTE_{xl_region_slug}_{fecha_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with d2:
        with st.spinner(""):
            orig_buf = build_original_format_excel(df_clean, df_consol, cols)
        st.download_button(
            label="📋 Base Limpia (formato original)",
            data=orig_buf,
            file_name=f"GASO_INOUT_LIMPIO_{fecha_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with d3:
        csv_buf = df_clean.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            label="📄 Base Limpia CSV",
            data=csv_buf,
            file_name=f"GASO_INOUT_LIMPIO_{fecha_str}.csv",
            mime="text/csv",
            use_container_width=True,
        )

    # Footer
    st.markdown(f"""
    <div style="text-align:center;padding:1.2rem;margin-top:1.5rem;
                border-top:1px solid #DCE8F5;color:#AAA;font-size:.75rem">
      GASO COMUNICACIONES · IN-OUT Report Processor v3.0 ·
      Generado el {datetime.date.today().strftime('%d/%m/%Y')}
    </div>
    """, unsafe_allow_html=True)
